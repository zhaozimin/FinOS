#!/usr/bin/env python3
import argparse
from collections import Counter, defaultdict
import hashlib
import json
from pathlib import Path
import shutil

from openpyxl import load_workbook

from finance_node_server import (
    CONFIG_PATH,
    DB_PATH,
    coerce_bool,
    coerce_float,
    connect_db,
    default_projects,
    ensure_schema,
    load_configuration_payload,
    load_config,
    normalize_keywords,
    normalize_projects,
    utc_now_iso,
)


def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text == "-" else text


def stable_id(prefix: str, *parts: str) -> str:
    digest = hashlib.sha1("||".join(parts).encode("utf-8")).hexdigest()[:12]
    return f"{prefix}-{digest}"


def workbook_settings(sheet) -> dict:
    settings = {}
    for key, value, *_ in sheet.iter_rows(min_row=2, values_only=True):
        if not key:
            continue
        settings[str(key)] = value
    return settings


def workbook_category_summary(sheet) -> dict:
    summary = {}
    for category_id, name, *_rest, keywords in sheet.iter_rows(min_row=2, values_only=True):
        category_name = clean_text(name)
        if not category_name:
            continue
        summary[category_name] = {
            "id": clean_text(category_id),
            "keywords": normalize_keywords(keywords, category_name),
        }
    return summary


def workbook_account_colors(sheet) -> dict:
    colors = {}
    for _account_id, name, *_rest, color in sheet.iter_rows(min_row=2, values_only=True):
        account_name = clean_text(name)
        if account_name:
            colors[account_name] = clean_text(color)
    return colors


def infer_ui_account_type(name: str) -> str:
    lowered = name.lower()
    if "储蓄" in name or "saving" in lowered:
        return "储蓄账户"
    if "投资" in name or "invest" in lowered:
        return "投资账户"
    if "应急" in name or "emergency" in lowered:
        return "应急账户"
    if any(token in name for token in ("生活", "微信", "支付宝")) or any(token in lowered for token in ("wechat", "alipay")):
        return "生活账户"
    if any(token in name for token in ("经营", "工资", "银行卡")) or any(token in lowered for token in ("company", "debit")):
        return "经营账户"
    return "其他"


def backend_account_type(ui_type: str) -> str:
    mapping = {
        "经营账户": "companyAccount",
        "生活账户": "digitalWallet",
        "投资账户": "investment",
        "应急账户": "emergencyFund",
        "储蓄账户": "savings",
    }
    return mapping.get(ui_type, "other")


def infer_logo_preset(name: str) -> str:
    lowered = name.lower()
    if "微信" in name or "wechat" in lowered:
        return "wechat-pay"
    if "支付宝" in name or "alipay" in lowered:
        return "alipay"
    if "招商" in name or "cmb" in lowered:
        return "cmb"
    return "virtual-card"


def infer_symbol_name(logo_preset_id: str) -> str:
    if logo_preset_id == "wechat-pay":
        return "message.fill"
    if logo_preset_id == "alipay":
        return "a.circle.fill"
    if logo_preset_id in {"cmb", "icbc", "abc", "boc", "ccb", "pingan"}:
        return "building.columns.fill"
    if logo_preset_id in {"paypal", "wise"}:
        return "globe"
    return "creditcard.fill"


def default_flow_role(ui_type: str) -> str:
    if ui_type == "经营账户":
        return "initial"
    if ui_type == "生活账户":
        return "both"
    return "internal"


def infer_category_group(name: str, direction: str) -> str:
    if direction == "收入":
        return "资金来源"
    if any(token in name for token in ("娱乐", "人情", "礼物", "聚会", "门票")):
        return "额外开销"
    if any(token in name for token in ("餐", "外卖", "购物", "住房", "医疗", "交通", "生活")):
        return "必要开销"
    return "公司运营"


def default_project_id_for_group(group: str) -> str:
    mapping = {
        "公司运营": "project-company-ops",
        "必要开销": "project-life-necessary",
        "额外开销": "project-life-extra",
    }
    return mapping.get(group, "project-company-ops")


def infer_category_icon(name: str, direction: str) -> str:
    if direction == "收入":
        return "banknote.fill"
    if "餐" in name or "外卖" in name:
        return "fork.knife"
    if "交通" in name:
        return "tram.fill"
    if "设备" in name:
        return "shippingbox.fill"
    if "娱乐" in name:
        return "gamecontroller.fill"
    if "红包" in name:
        return "gift.fill"
    if "办公" in name:
        return "briefcase.fill"
    return "tray"


def parse_transactions(sheet) -> list[dict]:
    transactions = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        raw_id, date_str, time_str, kind_label, title, amount, account_name, category_name, project_name, note, reimbursement_status, source = row
        title_text = clean_text(title)
        account_text = clean_text(account_name)
        category_text = clean_text(category_name)
        if not title_text or not account_text or not category_text or not date_str:
            continue

        occurred_at = f"{clean_text(date_str)}T{clean_text(time_str) or '00:00'}:00+08:00"
        kind = "收入" if clean_text(kind_label) == "收入" else "转账" if clean_text(kind_label) == "转账" else "支出"
        transaction_id = stable_id(
            "tx",
            clean_text(date_str),
            clean_text(time_str),
            kind,
            title_text,
            str(coerce_float(amount, 0.0)),
            account_text,
            category_text,
            clean_text(project_name),
            clean_text(note),
            clean_text(source),
            clean_text(raw_id),
        )

        transactions.append(
            {
                "id": transaction_id,
                "title": title_text,
                "amount": abs(coerce_float(amount, 0.0)),
                "kind": {"收入": "income", "转账": "transfer"}.get(kind, "expense"),
                "occurred_at": occurred_at,
                "account_name": account_text,
                "from_account_name": account_text if kind != "收入" else None,
                "to_account_name": account_text if kind == "收入" else None,
                "category_name": category_text,
                "project_name": clean_text(project_name),
                "merchant": title_text,
                "note": clean_text(note),
                "reimbursement_status": clean_text(reimbursement_status) or "notApplicable",
                "source": clean_text(source) or "imported",
                "source_name": "",
                "tags": [clean_text(project_name)] if clean_text(project_name) else [],
            }
        )

    transactions.sort(key=lambda item: (item["occurred_at"], item["id"]))
    return transactions


def current_configuration_snapshot() -> dict:
    configuration = load_configuration_payload()
    balances = {}
    for account in configuration["accounts"]:
        balances[account["name"]] = coerce_float(account.get("currentBalance"), coerce_float(account.get("openingBalance"), 0.0))
    return {
        "configuration": configuration,
        "balances": balances,
    }


def build_accounts(transactions: list[dict], current_snapshot: dict, settings: dict, account_colors: dict) -> list[dict]:
    existing_accounts = {
        account["name"]: dict(account)
        for account in current_snapshot["configuration"]["accounts"]
        if account.get("name")
    }
    desired_balances = dict(current_snapshot["balances"])
    account_order = [account["name"] for account in current_snapshot["configuration"]["accounts"] if account.get("name")]

    for transaction in transactions:
        account_name = transaction["account_name"]
        if account_name not in account_order:
            account_order.append(account_name)
        desired_balances.setdefault(account_name, 0.0)

    delta_by_account = defaultdict(float)
    for transaction in transactions:
        amount = coerce_float(transaction["amount"], 0.0)
        if transaction["kind"] == "income":
            delta_by_account[transaction["to_account_name"] or transaction["account_name"]] += amount
        elif transaction["kind"] == "expense":
            delta_by_account[transaction["from_account_name"] or transaction["account_name"]] -= amount
        elif transaction["kind"] == "transfer":
            delta_by_account[transaction["from_account_name"] or transaction["account_name"]] -= amount
            if transaction["to_account_name"]:
                delta_by_account[transaction["to_account_name"]] += amount

    accounts = []
    for index, name in enumerate(account_order):
        existing = existing_accounts.get(name, {})
        ui_type = clean_text(existing.get("uiAccountType")) or infer_ui_account_type(name)
        logo_preset_id = clean_text(existing.get("logoPresetId")) or infer_logo_preset(name)
        currency = clean_text(existing.get("currency")) or clean_text(settings.get("defaultCurrency")) or "CNY"
        desired_balance = coerce_float(desired_balances.get(name), 0.0)
        opening_balance = round(desired_balance - delta_by_account.get(name, 0.0), 2)
        account_id = clean_text(existing.get("id")) or stable_id("account", name)
        raw_threshold = existing.get("threshold")
        if raw_threshold in (None, ""):
            threshold = round(max(desired_balance, 0.0), 2)
        else:
            threshold = coerce_float(raw_threshold, round(max(desired_balance, 0.0), 2))

        accounts.append(
            {
                "id": account_id,
                "name": name,
                "type": clean_text(existing.get("type")) or backend_account_type(ui_type),
                "currency": currency,
                "openingBalance": opening_balance,
                "brand": clean_text(existing.get("brand")) or logo_preset_id,
                "tintHex": clean_text(existing.get("tintHex")) or account_colors.get(name) or "#607D8B",
                "symbolName": clean_text(existing.get("symbolName")) or infer_symbol_name(logo_preset_id),
                "keywords": normalize_keywords(existing.get("keywords"), name),
                "uiAccountType": ui_type,
                "customType": clean_text(existing.get("customType")),
                "logoMode": clean_text(existing.get("logoMode")) or "preset",
                "logoPresetId": logo_preset_id,
                "logoEmoji": clean_text(existing.get("logoEmoji")),
                "logoImageUrl": clean_text(existing.get("logoImageUrl")),
                "threshold": threshold,
                "note": clean_text(existing.get("note")),
                "flowRole": clean_text(existing.get("flowRole")) or default_flow_role(ui_type),
                "sortOrder": index,
            }
        )

    return accounts


def build_categories(transactions: list[dict], current_snapshot: dict, category_summary: dict, accounts: list[dict]) -> list[dict]:
    existing_categories = {
        category["name"]: dict(category)
        for category in current_snapshot["configuration"]["categories"]
        if category.get("name")
    }
    account_id_by_name = {account["name"]: account["id"] for account in accounts}
    category_stats = {}

    for transaction in transactions:
        category_name = transaction["category_name"]
        stats = category_stats.setdefault(
            category_name,
            {
                "direction": Counter(),
                "accounts": Counter(),
                "order": len(category_stats),
            },
        )
        stats["direction"]["收入" if transaction["kind"] == "income" else "支出"] += 1
        stats["accounts"][transaction["account_name"]] += 1

    categories = []
    for category_name, stats in sorted(category_stats.items(), key=lambda item: item[1]["order"]):
        existing = existing_categories.get(category_name, {})
        summary = category_summary.get(category_name, {})
        direction = existing.get("direction") if clean_text(existing.get("direction")) in {"收入", "支出"} else stats["direction"].most_common(1)[0][0]
        group = clean_text(existing.get("group")) or infer_category_group(category_name, direction)
        preferred_account_name = clean_text(existing.get("defaultAccountName")) or (stats["accounts"].most_common(1)[0][0] if stats["accounts"] else "")
        default_account_id = clean_text(existing.get("defaultAccountId")) or account_id_by_name.get(preferred_account_name, accounts[0]["id"])
        project_id = clean_text(existing.get("projectId")) or ("" if direction == "收入" else default_project_id_for_group(group))

        categories.append(
            {
                "id": clean_text(existing.get("id")) or clean_text(summary.get("id")) or stable_id("category", category_name),
                "name": category_name,
                "systemImage": clean_text(existing.get("systemImage")) or infer_category_icon(category_name, direction),
                "tintHex": clean_text(existing.get("tintHex")) or "#607D8B",
                "keywords": normalize_keywords(existing.get("keywords") or summary.get("keywords"), category_name),
                "direction": direction,
                "group": group,
                "defaultAccountId": default_account_id,
                "projectId": project_id,
                "note": clean_text(existing.get("note")),
                "sortOrder": len(categories),
            }
        )

    return categories


def build_settings(current_snapshot: dict, workbook_setting_values: dict) -> dict:
    current_settings = dict(current_snapshot["configuration"]["settings"])
    projects = normalize_projects(current_settings.get("projects"))
    return {
        "bookMode": clean_text(workbook_setting_values.get("bookMode")) or clean_text(current_settings.get("bookMode")) or "openClawDashboard",
        "defaultCurrency": clean_text(workbook_setting_values.get("defaultCurrency")) or clean_text(current_settings.get("defaultCurrency")) or "CNY",
        "baseUnit": clean_text(current_settings.get("baseUnit")) or "yuan",
        "timezone": clean_text(workbook_setting_values.get("timezone")) or clean_text(current_settings.get("timezone")) or "Asia/Shanghai",
        "allowManualEntry": coerce_bool(workbook_setting_values.get("allowManualEntry"), coerce_bool(current_settings.get("allowManualEntry"), False)),
        "projects": projects,
        "updatedAt": utc_now_iso(),
    }


def transaction_category_payloads(categories: list[dict]) -> dict:
    return {
        category["name"]: {
            "id": category["id"],
            "name": category["name"],
            "systemImage": category["systemImage"],
            "tintHex": category["tintHex"],
            "keywords": category["keywords"],
            "direction": category["direction"],
            "group": category["group"],
            "defaultAccountId": category["defaultAccountId"],
            "projectId": category["projectId"],
            "note": category["note"],
        }
        for category in categories
    }


def write_database(accounts: list[dict], categories: list[dict], settings: dict, transactions: list[dict], backup_path: Path) -> None:
    ensure_schema()
    shutil.copy2(DB_PATH, backup_path)

    category_lookup = transaction_category_payloads(categories)
    now = utc_now_iso()
    connection = connect_db()
    try:
        connection.execute("BEGIN")
        connection.execute("DELETE FROM transactions")
        connection.execute("DELETE FROM categories")
        connection.execute("DELETE FROM accounts")
        connection.execute("DELETE FROM ledger_settings")

        connection.executemany(
            """
            INSERT INTO categories (id, payload_json, sort_order, updated_at)
            VALUES (?, ?, ?, ?)
            """,
            [
                (
                    category["id"],
                    json.dumps({key: value for key, value in category.items() if key != "sortOrder"}, ensure_ascii=False),
                    category["sortOrder"],
                    now,
                )
                for category in categories
            ],
        )

        connection.executemany(
            """
            INSERT INTO accounts (id, payload_json, sort_order, updated_at)
            VALUES (?, ?, ?, ?)
            """,
            [
                (
                    account["id"],
                    json.dumps({key: value for key, value in account.items() if key != "sortOrder"}, ensure_ascii=False),
                    account["sortOrder"],
                    now,
                )
                for account in accounts
            ],
        )

        connection.execute(
            """
            INSERT INTO ledger_settings (id, payload_json, updated_at)
            VALUES (1, ?, ?)
            """,
            (json.dumps(settings, ensure_ascii=False), now),
        )

        connection.executemany(
            """
            INSERT INTO transactions (
                id, title, amount, kind, occurred_at, category_json, tags_json,
                account_name, from_account_name, to_account_name, merchant, project_name,
                note, reimbursement_status, source, source_name, created_at, updated_at
            ) VALUES (
                :id, :title, :amount, :kind, :occurred_at, :category_json, :tags_json,
                :account_name, :from_account_name, :to_account_name, :merchant, :project_name,
                :note, :reimbursement_status, :source, :source_name, :created_at, :updated_at
            )
            """,
            [
                {
                    **transaction,
                    "category_json": json.dumps(category_lookup[transaction["category_name"]], ensure_ascii=False),
                    "tags_json": json.dumps(transaction["tags"], ensure_ascii=False),
                    "created_at": now,
                    "updated_at": now,
                }
                for transaction in transactions
            ],
        )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def update_runtime_config() -> None:
    config = load_config()
    config["lastIngestedAt"] = utc_now_iso()
    with CONFIG_PATH.open("w", encoding="utf-8") as file:
        json.dump(config, file, ensure_ascii=False, indent=2)


def main() -> None:
    parser = argparse.ArgumentParser(description="Import a corrected finance workbook into Finance Node.")
    parser.add_argument("workbook", type=Path, help="Path to the exported Excel workbook")
    parser.add_argument("--backup", type=Path, default=None, help="Optional backup path for finance.sqlite3")
    args = parser.parse_args()

    workbook_path = args.workbook.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)
    transactions = parse_transactions(workbook["交易流水"])
    current_snapshot = current_configuration_snapshot()
    settings = build_settings(current_snapshot, workbook_settings(workbook["系统设置"]))
    account_colors = workbook_account_colors(workbook["账户余额"])
    accounts = build_accounts(transactions, current_snapshot, settings, account_colors)
    categories = build_categories(transactions, current_snapshot, workbook_category_summary(workbook["分类汇总"]), accounts)

    backup_path = args.backup or DB_PATH.with_suffix(f".backup-{datetime_suffix()}.sqlite3")
    write_database(accounts, categories, settings, transactions, backup_path)
    update_runtime_config()

    print(f"Imported {len(transactions)} transactions into {DB_PATH}")
    print(f"Backed up previous database to {backup_path}")
    print(f"Accounts: {len(accounts)} | Categories: {len(categories)}")


def datetime_suffix() -> str:
    return utc_now_iso().replace(":", "").replace("+", "_plus_")


if __name__ == "__main__":
    main()
