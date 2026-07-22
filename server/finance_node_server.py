#!/usr/bin/env python3
"""
[INPUT]: 依赖 runtime SQLite、静态 PWA 产物和 agent_audit 的审计/检查点能力。
[OUTPUT]: 对外提供 Finance Node HTTP API、主数据 Agent 操作与可见删除状态。
[POS]: service 的应用入口；协调账本持久化、鉴权、Agent 工具和前端数据契约。
[PROTOCOL]: 变更时更新此头部，然后检查 CLAUDE.md
"""
from collections import defaultdict
from contextlib import closing
import base64
import gzip
import hmac
import json
import mimetypes
import os
import sqlite3
import sys
import threading
from datetime import datetime, timezone
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Optional, Union
from urllib.parse import parse_qs, urlparse
from uuid import uuid4

from agent_audit import append_audit_event, checkpoint_database, ensure_audit_schema

mimetypes.add_type("application/manifest+json", ".webmanifest")

ROOT = Path(__file__).resolve().parent
RUNTIME_DIR = ROOT / "runtime"
CONFIG_PATH = RUNTIME_DIR / "config.json"
DB_PATH = RUNTIME_DIR / "finance.sqlite3"
ATTACHMENTS_DIR = RUNTIME_DIR / "attachments"
ATTACHMENT_MAX_BYTES = 10 * 1024 * 1024  # 10 MB hard cap per file
# 请求体硬顶：base64 后的 10MB 附件约 13.3MB + JSON 包裹，取 32MB 覆盖合法导入/上传，
# 读入前即拒绝超限，防超大 body 内存耗尽 DoS。
MAX_JSON_BODY_BYTES = 32 * 1024 * 1024
WEB_DIR = ROOT / "web"
# 静态资源的唯一可读根：任何 _send_file 目标必须解析后仍落在此目录内，
# 否则视为目录穿越（读 runtime/config.json 偷 token、下载整库、读任意文件）。
WEB_ROOT = WEB_DIR.resolve()
MUTATION_LOCK = threading.RLock()


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalize_iso8601(value: Optional[str], fallback: Optional[str] = None) -> str:
    if not value:
        return fallback or utc_now_iso()
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).isoformat()
    except ValueError:
        return fallback or utc_now_iso()


def load_config() -> dict:
    if not CONFIG_PATH.exists():
        raise RuntimeError(f"Missing config file: {CONFIG_PATH}")
    with CONFIG_PATH.open("r", encoding="utf-8") as file:
        return json.load(file)


def connect_db() -> sqlite3.Connection:
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def default_categories() -> list:
    # 新手默认：7 支出 + 3 收入 + 1 系统（转账，前端从可编辑列表隐藏）。
    # 每个类别都显式带 direction；收入类预挂 defaultAccountId，一装即可用。
    return [
        {
            "id": "category-food",
            "name": "餐饮",
            "systemImage": "fork.knife",
            "tintHex": "#FF8C42",
            "direction": "支出",
            "keywords": ["饭", "午饭", "晚饭", "咖啡", "奶茶", "早餐", "餐", "火锅"],
        },
        {
            "id": "category-transport",
            "name": "交通",
            "systemImage": "tram.fill",
            "tintHex": "#4F7CFF",
            "direction": "支出",
            "keywords": ["打车", "地铁", "高铁", "机票", "机场", "滴滴", "车费"],
        },
        {
            "id": "category-housing",
            "name": "住房",
            "systemImage": "house.fill",
            "tintHex": "#6F5BD3",
            "direction": "支出",
            "keywords": ["房租", "物业", "水电", "酒店", "住宿"],
        },
        {
            "id": "category-shopping",
            "name": "购物",
            "systemImage": "bag.fill",
            "tintHex": "#E64980",
            "direction": "支出",
            "keywords": ["购物", "淘宝", "衣服", "鞋", "日用", "京东"],
        },
        {
            "id": "category-entertainment",
            "name": "娱乐",
            "systemImage": "gamecontroller.fill",
            "tintHex": "#F06543",
            "direction": "支出",
            "keywords": ["电影", "游戏", "聚餐", "娱乐", "门票"],
        },
        {
            "id": "category-health",
            "name": "医疗",
            "systemImage": "cross.case.fill",
            "tintHex": "#FF5D73",
            "direction": "支出",
            "keywords": ["医院", "药", "体检", "挂号"],
        },
        {
            "id": "category-social",
            "name": "人情往来",
            "systemImage": "gift.fill",
            "tintHex": "#E8590C",
            "direction": "支出",
            "keywords": ["红包", "礼金", "份子", "请客", "送礼", "人情"],
        },
        {
            "id": "category-salary",
            "name": "工资",
            "systemImage": "banknote.fill",
            "tintHex": "#2F9E44",
            "direction": "收入",
            "defaultAccountId": "account-salary",
            "keywords": ["工资", "薪水", "月薪", "发工资", "奖金", "年终奖"],
        },
        {
            "id": "category-redpacket",
            "name": "红包",
            "systemImage": "gift.fill",
            "tintHex": "#37B24D",
            "direction": "收入",
            "defaultAccountId": "account-wechat",
            "keywords": ["红包", "收红包", "转账收款", "收款"],
        },
        {
            "id": "category-other-income",
            "name": "其他收入",
            "systemImage": "plus.circle.fill",
            "tintHex": "#1C7ED6",
            "direction": "收入",
            "keywords": ["收入", "报销到账", "退款", "利息", "返现"],
        },
        {
            "id": "category-transfer",
            "name": "转账",
            "systemImage": "arrow.left.arrow.right.square.fill",
            "tintHex": "#546E7A",
            "direction": "支出",
            "keywords": ["转账", "转入", "转出"],
        },
    ]


def default_accounts() -> list:
    return [
        {
            "id": "account-wechat",
            "name": "微信支付",
            "type": "digitalWallet",
            "currency": "CNY",
            "openingBalance": 0.0,
            "threshold": 0.0,
            "brand": "wechat",
            "tintHex": "#07C160",
            "symbolName": "message.fill",
            "keywords": ["微信", "wechat"],
            "ownership": "personal",
        },
        {
            "id": "account-alipay",
            "name": "支付宝",
            "type": "digitalWallet",
            "currency": "CNY",
            "openingBalance": 0.0,
            "threshold": 0.0,
            "brand": "alipay",
            "tintHex": "#1677FF",
            "symbolName": "qrcode",
            "keywords": ["支付宝", "alipay"],
            "ownership": "personal",
        },
        {
            # 用角色名而非品牌名：新手进来就知道往哪填，装完自行改成自己的银行。
            "id": "account-salary",
            "name": "工资卡",
            "type": "debitCard",
            "currency": "CNY",
            "openingBalance": 0.0,
            "threshold": 0.0,
            "brand": "custom",
            "tintHex": "#5C6BC0",
            "symbolName": "building.columns.fill",
            "keywords": ["工资", "工资卡", "银行卡", "储蓄卡", "借记卡"],
            "ownership": "personal",
        },
    ]


VALID_OWNERSHIPS = {"company", "personal", "unspecified"}


def normalize_ownership(value: object, fallback: str = "unspecified") -> str:
    """Normalize an ownership value to one of company / personal / unspecified."""
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in VALID_OWNERSHIPS:
            return lowered
    return fallback if fallback in VALID_OWNERSHIPS else "unspecified"


def infer_account_ownership(payload: dict) -> str:
    """Infer ownership for an account that lacks an explicit ownership field.

    Rules (Chinese-first market):
        company   ← uiAccountType "经营账户" / name 含 "经营"|"公司"|"工资"
        personal  ← uiAccountType "生活账户"|"应急账户"|"储蓄账户" / name 含 "生活"|"应急"|"储蓄"|"微信"|"支付宝"
        unspecified ← otherwise (including 投资账户)
    """
    explicit = payload.get("ownership")
    if isinstance(explicit, str) and explicit.strip().lower() in VALID_OWNERSHIPS:
        return explicit.strip().lower()

    ui_type = str(payload.get("uiAccountType") or "").strip()
    name = str(payload.get("name") or "").strip()
    name_lower = name.lower()

    if ui_type == "经营账户" or any(token in name for token in ("经营", "公司", "工资")):
        return "company"
    if ui_type in {"生活账户", "应急账户", "储蓄账户"}:
        return "personal"
    if any(token in name for token in ("生活", "应急", "储蓄", "微信", "支付宝")):
        return "personal"
    if "wechat" in name_lower or "alipay" in name_lower:
        return "personal"
    return "unspecified"


def default_exchange_rates() -> dict:
    """1.7 多币种：默认汇率（仅供参考；用户在设置页手动覆盖）。
    单位：每 1 单位外币兑换的 CNY 量（baseCurrency=CNY 时 CNY=1）。"""
    return {
        "baseCurrency": "CNY",
        "rates": {
            "CNY": 1.0,
            "USD": 7.20,
            "HKD": 0.92,
            "EUR": 7.80,
            "JPY": 0.048,
        },
        "autoFetch": False,
        "provider": "open.er-api.com",
        "lastFetchSource": None,
        "lastFetchError": None,
        "updatedAt": None,
    }


def normalize_exchange_rates(value: object) -> dict:
    base = default_exchange_rates()
    if isinstance(value, dict):
        if isinstance(value.get("baseCurrency"), str):
            base["baseCurrency"] = value["baseCurrency"].upper().strip() or base["baseCurrency"]
        if isinstance(value.get("rates"), dict):
            normalized: dict[str, float] = {}
            for code, rate in value["rates"].items():
                if not isinstance(code, str):
                    continue
                code_clean = code.upper().strip()
                if not code_clean:
                    continue
                normalized[code_clean] = max(0.0, coerce_float(rate, 0.0))
            # base currency 必须是 1（不允许编辑掉）
            normalized[base["baseCurrency"]] = 1.0
            base["rates"] = normalized or base["rates"]
        if "autoFetch" in value:
            base["autoFetch"] = coerce_bool(value.get("autoFetch"), False)
        if isinstance(value.get("provider"), str):
            base["provider"] = value["provider"].strip() or base["provider"]
        if isinstance(value.get("lastFetchSource"), str):
            base["lastFetchSource"] = value["lastFetchSource"]
        if isinstance(value.get("lastFetchError"), str):
            base["lastFetchError"] = value["lastFetchError"]
        if isinstance(value.get("updatedAt"), str):
            base["updatedAt"] = value["updatedAt"]
    return base


def fetch_exchange_rates_from_api(base_currency: str = "CNY") -> tuple[dict, Optional[str]]:
    """从 open.er-api.com 拉取 base 币种对其他币种的汇率。
    返回 (rates_dict, error_message)。
    rates_dict 的格式与 LedgerSettings.exchangeRates.rates 一致：1 X = N base。
    """
    import urllib.request
    import urllib.error

    base = (base_currency or "CNY").upper()
    url = f"https://open.er-api.com/v6/latest/{base}"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "FinanceNode/0.1"})
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.URLError as exc:
        return ({}, f"网络错误：{exc.reason}")
    except (json.JSONDecodeError, TimeoutError) as exc:
        return ({}, f"响应解析失败：{exc}")
    except Exception as exc:  # pragma: no cover
        return ({}, f"未知错误：{exc}")

    if data.get("result") != "success":
        return ({}, f"API 返回失败：{data.get('error-type', 'unknown')}")
    raw_rates = data.get("rates") or {}
    # API 返回的是 "1 base = N other"。我们存储格式是 "1 other = N base"，所以反转。
    out: dict[str, float] = {base: 1.0}
    for code, value in raw_rates.items():
        try:
            v = float(value)
        except (TypeError, ValueError):
            continue
        if v <= 0:
            continue
        out[code.upper()] = round(1.0 / v, 6)
    return (out, None)


def refresh_exchange_rates_in_settings() -> dict:
    """读 ledger_settings.exchangeRates，调 fetch_exchange_rates_from_api，
    把结果合并回 settings（保留用户手动加的小币种），写回 DB。
    返回更新后的 exchangeRates dict（含 lastFetchSource / lastFetchError / updatedAt）。"""
    connection = connect_db()
    try:
        settings = load_ledger_settings(connection)
        rates_cfg = normalize_exchange_rates(settings.get("exchangeRates"))
        base = rates_cfg.get("baseCurrency", "CNY")
        new_rates, err = fetch_exchange_rates_from_api(base)
        now = utc_now_iso()
        if err:
            rates_cfg["lastFetchError"] = err
        else:
            # 合并：自动拉到的覆盖，用户手动加的小币种保留
            merged = dict(rates_cfg.get("rates") or {})
            merged.update(new_rates)
            merged[base] = 1.0  # 强制 base = 1
            rates_cfg["rates"] = merged
            rates_cfg["lastFetchSource"] = rates_cfg.get("provider") or "open.er-api.com"
            rates_cfg["lastFetchError"] = None
            rates_cfg["updatedAt"] = now
        settings["exchangeRates"] = rates_cfg
        settings["updatedAt"] = now
        connection.execute(
            "UPDATE ledger_settings SET payload_json = ?, updated_at = ? WHERE id = 1",
            (json.dumps(settings, ensure_ascii=False), now),
        )
        connection.commit()
        return rates_cfg
    finally:
        connection.close()


def default_tax_config() -> dict:
    """2.3 税务配置默认值（中国小规模 / 个体户常见参数，可在设置页改）。"""
    return {
        "vatRate": 0.03,                # 增值税率 3%（小规模简易计税）
        "personalThreshold": 60000,     # 个税起征点 6 万 / 年
        "personalRate": 0.20,           # 简化个税率（实际累进，先用单率近似）
        "sebRate": 0.10,                # 社保 + 公积金合计费率（10%，可改）
        "currency": "CNY",
        "note": "本数据仅供参考，请以专业税务人员意见为准。",
    }


def normalize_tax_config(value: object) -> dict:
    base = default_tax_config()
    if isinstance(value, dict):
        for key in ("vatRate", "personalRate", "sebRate"):
            if key in value:
                base[key] = max(0.0, coerce_float(value.get(key), base[key]))
        if "personalThreshold" in value:
            base["personalThreshold"] = max(0.0, coerce_float(value.get("personalThreshold"), base["personalThreshold"]))
        if "currency" in value:
            base["currency"] = str(value.get("currency") or base["currency"])
        if "note" in value:
            base["note"] = str(value.get("note") or base["note"])
    return base


def default_ledger_settings() -> dict:
    return {
        "bookMode": "personalAssistant",
        # 记账模式：新装默认 personal（个人），隐藏归属/经营/税务等创业者维度；
        # dual = 个人 + 经营（今日全量形态）。老库由 infer_ledger_mode 迁移推断。
        "ledgerMode": "personal",
        "defaultCurrency": "CNY",
        "baseUnit": "yuan",
        "timezone": "Asia/Shanghai",
        "allowManualEntry": True,
        "projects": default_projects(),
        "financeSources": default_finance_sources(),
        "counterparties": [],
        "taxConfig": default_tax_config(),
        "exchangeRates": default_exchange_rates(),
        "updatedAt": None,
    }


VALID_LEDGER_MODES = {"personal", "dual"}


def normalize_ledger_mode(value: object, fallback: str = "personal") -> str:
    """记账模式归一化：personal（个人）| dual（个人 + 经营）。"""
    if isinstance(value, str) and value.strip().lower() in VALID_LEDGER_MODES:
        return value.strip().lower()
    return fallback if fallback in VALID_LEDGER_MODES else "personal"


def infer_ledger_mode(connection: sqlite3.Connection) -> str:
    """老库迁移：payload 无 ledgerMode 字段时，存在 company 归属账户即判为 dual。"""
    try:
        for row in connection.execute("SELECT payload_json FROM accounts").fetchall():
            try:
                payload = json.loads(row["payload_json"])
            except (json.JSONDecodeError, TypeError):
                continue
            if isinstance(payload, dict) and payload.get("ownership") == "company":
                return "dual"
    except sqlite3.Error:
        return "personal"
    return "personal"


def scrub_account_ref(account_id: object, valid_account_ids: set) -> str:
    """引用完整性：defaultAccountId 指向不存在的账户时清空，杜绝悬空引用。"""
    aid = str(account_id or "")
    return aid if aid in valid_account_ids else ""


# ============================================================
# AI 主数据防护闸（2026-07）
# 原则：AI 只可建议，用户确认才落库。
#  - openClaw 通道的交易引用未知账户/类别/来源/项目 → 422 + 全部合法选项，
#    逼 Agent 回到对话里让用户从「新建 / 改用现有 / 放弃」中拍板；
#  - /v1/agent/operations 的主数据增删改必须携带 userConfirmation（用户确认原话），
#    并随审计事件落库存证。人类 UI 通道（dashboard/manual/import）不受影响。
# ============================================================

# 人工/系统写入通道：仅这些 source 的交易豁免 AI 主数据护栏（它们只从 UI 下拉/
# 系统流程引用已存在的主数据）。其余一切来源——agent / claude / gpt / openclaw /
# 缺省 openClaw 等——都必须过引用校验（先问后写）。反转白名单，杜绝换个 source 就绕过。
HUMAN_OR_SYSTEM_SOURCES = {"manual", "dashboard", "adjustment", "recurring"}


def source_requires_guardrail(source: object) -> bool:
    s = str(source or "").strip().lower()
    if not s:
        return True  # 缺省即当作 agent，强制校验
    if s == "import" or s.startswith("import-"):
        return False  # 账单导入是系统批量写入
    return s not in HUMAN_OR_SYSTEM_SOURCES

# 内部流程写入、不属于用户类别配置的系统类别（如余额调整审计交易）
INTERNAL_CATEGORY_NAMES = {"余额调整"}


def collect_master_registry(connection: sqlite3.Connection) -> dict:
    """汇总当前有效（未删除）主数据名录，供引用校验与 422 提示复用。"""
    accounts = [str(a.get("name")) for a in list_accounts(connection) if a.get("name") and not a.get("deletedAt")]
    categories = [str(c.get("name")) for c in list_categories(connection) if c.get("name") and not c.get("deletedAt")]
    settings = load_ledger_settings(connection)
    sources = [str(s.get("name")) for s in settings.get("financeSources", []) if s.get("name") and not s.get("deletedAt")]
    projects = [str(p.get("name")) for p in settings.get("projects", []) if p.get("name")]
    return {"accounts": accounts, "categories": categories, "sources": sources, "projects": projects}


def _category_name_from_row(row: object) -> str:
    try:
        return str(json.loads(row["category_json"] or "{}").get("name") or "").strip()
    except (json.JSONDecodeError, TypeError, KeyError, IndexError):
        return ""


def validate_agent_transaction_refs(
    connection: sqlite3.Connection, row: dict, existing_row: Optional[sqlite3.Row] = None
) -> Optional[dict]:
    """openClaw 交易的引用校验；PUT 更新只查相对原记录变化的字段，不误伤存量编辑。
    返回 None = 通过；返回 dict = 应以 422 返回的结构化错误。"""
    if not source_requires_guardrail(row.get("source")):
        return None

    registry = collect_master_registry(connection)
    issues: list[dict] = []

    def changed(key: str) -> bool:
        if existing_row is None:
            return True
        try:
            return str(row.get(key) or "") != str(existing_row[key] or "")
        except (KeyError, IndexError):
            return True

    def check(kind: str, field: str, value: object, valid: set) -> None:
        text = str(value or "").strip()
        if text and text not in valid:
            issues.append({"field": field, "value": text, "kind": kind})

    account_set = set(registry["accounts"])
    if changed("account_name"):
        check("account", "accountName", row.get("account_name"), account_set)
    if changed("from_account_name"):
        check("account", "fromAccountName", row.get("from_account_name"), account_set)
    if changed("to_account_name"):
        check("account", "toAccountName", row.get("to_account_name"), account_set)

    category_name = _category_name_from_row(row)
    previous_category = _category_name_from_row(existing_row) if existing_row is not None else None
    if (
        category_name
        and category_name != previous_category
        and category_name not in INTERNAL_CATEGORY_NAMES
        and category_name not in set(registry["categories"])
    ):
        issues.append({"field": "category", "value": category_name, "kind": "category"})

    if changed("source_name"):
        check("financeSource", "sourceName", row.get("source_name"), set(registry["sources"]))
    if changed("project_name"):
        check("project", "projectName", row.get("project_name"), set(registry["projects"]))

    if not issues:
        return None

    kind_label = {"account": "账户", "category": "类别", "financeSource": "资金来源", "project": "项目"}
    detail = "、".join(f"{kind_label[item['kind']]}「{item['value']}」" for item in issues)
    return {
        "error": f"{detail} 不存在。AI 不得静默创建或修改主数据——请先向用户确认。",
        "code": "unknown_master_data",
        "issues": issues,
        "valid": registry,
        "agentInstruction": (
            "向用户逐项列出选择并等待明确答复，禁止代替用户决定："
            "1) 新建该主数据（用户同意后调用主数据工具，userConfirmation 填用户确认原话）；"
            "2) 改用 valid 列表中的现有项（把候选报给用户挑）；"
            "3) 放弃本笔记录。确认后重发本请求。"
        ),
    }


def default_finance_sources() -> list:
    # 新手默认：两个最常见的收入来源，各自预挂到真实存在的账户，杜绝满屏"不指定"。
    return [
        {
            "id": "source-salary",
            "name": "工资",
            "defaultAccountId": "account-salary",
            "note": "月薪 / 劳务，默认入工资卡。",
            "tintHex": "#5C6BC0",
        },
        {
            "id": "source-redpacket",
            "name": "红包",
            "defaultAccountId": "account-wechat",
            "note": "微信红包 / 转账收款。",
            "tintHex": "#37B24D",
        },
        {
            "id": "source-reimbursement",
            "name": "报销回款",
            "defaultAccountId": "account-salary",
            "note": "报销打回来的钱，默认入工资卡；到账后把原支出的报销状态改为已报销。",
            "tintHex": "#E8A33D",
        },
    ]


def default_projects() -> list:
    # 新手默认：只留个人侧两档（必要/额外）。经营项目属 dual 模式，由用户自行添加。
    return [
        {
            "id": "project-life-necessary",
            "name": "必要开销",
            "direction": "支出",
            "group": "必要开销",
            "note": "固定且必要的生活支出。",
            "trackingEnabled": False,
        },
        {
            "id": "project-life-extra",
            "name": "额外开销",
            "direction": "支出",
            "group": "额外开销",
            "note": "弹性和可选的生活消费。",
            "trackingEnabled": False,
        },
    ]


def coerce_float(value: object, default: float = 0.0) -> float:
    try:
        if value in {None, ""}:
            return default
        result = float(value)
    except (TypeError, ValueError):
        return default
    # 拒绝 NaN/Infinity（JSON 默认 allow_nan，会污染汇总/余额），回落默认值。
    if result != result or result in (float("inf"), float("-inf")):
        return default
    return result


def coerce_bool(value: object, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"1", "true", "yes", "on"}:
            return True
        if lowered in {"0", "false", "no", "off"}:
            return False
    return bool(value)


def normalize_keywords(value: object, fallback: Optional[str] = None) -> list[str]:
    if isinstance(value, list):
        keywords = [str(item).strip() for item in value if str(item).strip()]
        return keywords or ([fallback] if fallback else [])
    if isinstance(value, str):
        keywords = [item.strip() for item in value.split(",") if item.strip()]
        return keywords or ([fallback] if fallback else [])
    return [fallback] if fallback else []


def normalize_projects(items: object) -> list[dict]:
    if not isinstance(items, list):
        return default_projects()

    normalized = []
    for index, item in enumerate(items):
        if not isinstance(item, dict):
            continue
        project_id = item.get("id") or f"project-{index + 1}"
        name = str(item.get("name") or f"项目 {index + 1}").strip() or f"项目 {index + 1}"
        direction = "收入" if item.get("direction") == "收入" else "支出"
        group = str(item.get("group") or name).strip() or name
        # 1.3 储蓄目标：保留 goal 子对象。targetAmount 必填且 > 0 才算"有目标"
        goal_raw = item.get("goal")
        goal = None
        if isinstance(goal_raw, dict):
            target_amount = coerce_float(goal_raw.get("targetAmount"), 0.0)
            if target_amount > 0:
                goal = {
                    "targetAmount": target_amount,
                    "targetDate": str(goal_raw.get("targetDate") or "")[:10] or None,
                    "sourceAccountId": str(goal_raw.get("sourceAccountId") or ""),
                    "description": str(goal_raw.get("description") or ""),
                }
        normalized.append(
            {
                "id": project_id,
                "name": name,
                "direction": direction,
                "group": group,
                "note": str(item.get("note") or ""),
                "trackingEnabled": coerce_bool(item.get("trackingEnabled"), False),
                "goal": goal,
                # 2.1 项目预算 / 期望收入
                "expectedCost": coerce_float(item.get("expectedCost"), 0.0),
                "expectedRevenue": coerce_float(item.get("expectedRevenue"), 0.0),
                "startDate": str(item.get("startDate") or "")[:10] or None,
                "endDate": str(item.get("endDate") or "")[:10] or None,
            }
        )

    return normalized or default_projects()


def normalize_finance_sources(items: object) -> list[dict]:
    if not isinstance(items, list):
        return default_finance_sources()

    normalized = []
    for index, item in enumerate(items):
        if not isinstance(item, dict):
            continue
        source_id = item.get("id") or f"source-{index + 1}"
        name = str(item.get("name") or f"资金来源 {index + 1}").strip() or f"资金来源 {index + 1}"
        normalized.append(
            {
                "id": source_id,
                "name": name,
                "defaultAccountId": str(item.get("defaultAccountId") or ""),
                "note": str(item.get("note") or ""),
                "tintHex": str(item.get("tintHex") or "#87B99B"),
                "deletedAt": item.get("deletedAt") or None,
                "deletedBy": item.get("deletedBy") or None,
                "deletionReason": item.get("deletionReason") or None,
            }
        )

    return normalized or default_finance_sources()


_COUNTERPARTY_KINDS = ("client", "vendor", "employer", "other")


def normalize_counterparties(items: object) -> list[dict]:
    """2.4 客户 / 合作方名册。空列表是合理的默认值。"""
    if not isinstance(items, list):
        return []
    normalized = []
    for index, item in enumerate(items):
        if not isinstance(item, dict):
            continue
        cid = item.get("id") or f"counterparty-{index + 1}"
        name = str(item.get("name") or f"对手方 {index + 1}").strip() or f"对手方 {index + 1}"
        kind = item.get("kind") if item.get("kind") in _COUNTERPARTY_KINDS else "client"
        normalized.append(
            {
                "id": cid,
                "name": name,
                "kind": kind,
                "tintHex": str(item.get("tintHex") or "#7F91D6"),
                "defaultAccountId": str(item.get("defaultAccountId") or ""),
                "note": str(item.get("note") or ""),
                "contactInfo": str(item.get("contactInfo") or ""),
            }
        )
    return normalized


_RECURRING_FREQUENCIES = ("daily", "weekly", "monthly", "yearly")


def normalize_recurring_payload(item: dict) -> dict:
    """1.2 周期性交易：把前端 payload 标准化成可写库的字段（不含 id / created_at）。"""
    name = str(item.get("name") or "").strip() or "周期账目"
    template = item.get("template") if isinstance(item.get("template"), dict) else {}
    frequency = item.get("frequency") if item.get("frequency") in _RECURRING_FREQUENCIES else "monthly"
    interval_n = max(1, int(coerce_float(item.get("intervalN"), 1)))
    day_of_period_raw = item.get("dayOfPeriod")
    day_of_period = None
    if day_of_period_raw not in {None, "", 0}:
        try:
            day_of_period = int(day_of_period_raw)
        except (TypeError, ValueError):
            day_of_period = None
    today = datetime.now().astimezone().date().isoformat()
    start_date = str(item.get("startDate") or today)[:10]
    end_date_raw = item.get("endDate")
    end_date = str(end_date_raw)[:10] if end_date_raw else None
    next_due_at = str(item.get("nextDueAt") or start_date)[:10]
    enabled = 1 if coerce_bool(item.get("enabled"), True) else 0
    return {
        "name": name,
        "template_payload_json": json.dumps(template, ensure_ascii=False),
        "frequency": frequency,
        "interval_n": interval_n,
        "day_of_period": day_of_period,
        "start_date": start_date,
        "end_date": end_date,
        "next_due_at": next_due_at,
        "enabled": enabled,
    }


def recurring_row_to_dict(row: sqlite3.Row) -> dict:
    try:
        template = json.loads(row["template_payload_json"]) if row["template_payload_json"] else {}
    except (json.JSONDecodeError, TypeError):
        template = {}
    return {
        "id": row["id"],
        "name": row["name"],
        "template": template,
        "frequency": row["frequency"],
        "intervalN": int(row["interval_n"] or 1),
        "dayOfPeriod": row["day_of_period"],
        "startDate": row["start_date"],
        "endDate": row["end_date"],
        "nextDueAt": row["next_due_at"],
        "lastRunAt": row["last_run_at"],
        "enabled": bool(row["enabled"]),
        "createdAt": row["created_at"],
        "updatedAt": row["updated_at"],
    }


def advance_due_date(date_str: str, frequency: str, interval_n: int) -> str:
    """计算下次触发日期。月/年按"同一日"前进；不足则取月末。"""
    base = datetime.fromisoformat(date_str).date()
    interval_n = max(1, interval_n)
    if frequency == "daily":
        from datetime import timedelta
        return (base + timedelta(days=interval_n)).isoformat()
    if frequency == "weekly":
        from datetime import timedelta
        return (base + timedelta(weeks=interval_n)).isoformat()
    if frequency == "monthly":
        year = base.year
        month = base.month + interval_n
        while month > 12:
            month -= 12
            year += 1
        # 处理 2 月 / 月末日不存在
        from calendar import monthrange
        day = min(base.day, monthrange(year, month)[1])
        return f"{year:04d}-{month:02d}-{day:02d}"
    if frequency == "yearly":
        from calendar import monthrange
        year = base.year + interval_n
        day = min(base.day, monthrange(year, base.month)[1])
        return f"{year:04d}-{base.month:02d}-{day:02d}"
    # fallback: 按天前进
    from datetime import timedelta
    return (base + timedelta(days=1)).isoformat()


_RECURRING_CATCHUP_LAST_RUN: list = [None]
_RECURRING_CATCHUP_DEBOUNCE_SEC = 30


# ============== 1.4 账单导入：解析器模块 ==============
# 设计：
#   - parse_import_content(template, raw_bytes) -> {transactions: [...], warnings: [...]}
#   - 各 template（wechat / alipay / generic）共享 CSV 通用 parser，
#     模板差异主要在「跳过前 N 行 + 列名别名」。
#   - 所有解析后的字段都返回 dict，但不写库。前端可逐条编辑 / 勾选后再 commit。

import csv
import io
from datetime import timedelta as _timedelta

# 列名别名表：把各家原始列名映射到我们的通用字段。
# 这里覆盖了微信支付 / 支付宝 / 招行 / 中行 / 工行 / 通用 CSV 的常见列名。
_IMPORT_COLUMN_ALIASES = {
    # 时间
    "occurredAt": [
        "交易时间", "日期", "交易日期", "时间", "记账日期", "操作时间",
        "交易日", "交易时间(北京时间)", "记账日", "transaction time", "date", "occurredat",
        "trans date", "trade date",
    ],
    # 金额
    "amount": [
        "金额", "金额(元)", "支出金额", "收入金额", "金额（元）", "交易金额", "交易金额(元)",
        "金额（收入为正,支出为负）", "amount", "amt", "支出/收入", "收/支金额",
        "trans amount", "trade amount",
    ],
    # 类型 / 收支
    "kindLabel": [
        "收/支", "收支", "收入/支出", "type", "direction", "交易类型", "借贷标志", "借贷",
        "收支类型", "trans type",
    ],
    # 对方 / 商户
    "merchant": [
        "交易对方", "商户名称", "对方名称", "商品", "商品说明", "摘要",
        "对方账户名称", "对方户名", "对手户名", "对方账户", "merchant", "name", "payee",
        "交易摘要", "用途",
    ],
    # 备注
    "note": ["备注", "remark", "note", "memo", "附言", "用途备注"],
    # 分类
    "category": ["分类", "交易分类", "category", "支出分类", "类型"],
    # 支付方式 / 账户
    "accountName": [
        "支付方式", "账户", "支付账户", "card", "account", "支付方式 / 账户",
        "本方账户", "我的账户", "卡号",
    ],
}

# 用于识别"收入 / 支出"列的常见取值
_KIND_INCOME_TOKENS = {"收入", "income", "credit", "+", "贷", "进账", "存入", "已收"}
_KIND_EXPENSE_TOKENS = {"支出", "expense", "debit", "-", "借", "出账", "支取", "已付"}
_KIND_TRANSFER_TOKENS = {"转账", "transfer", "划转", "内部转账"}

# 各模板的预跳过行数（账单文件常带头部说明）
# 实际值通过 "找第一个像表头的行" 来动态校正（见 parse_import_content）。
# 这里只是兜底默认。
_TEMPLATE_SKIP_LINES = {
    "wechat": 16,   # 微信账单 ~16 行说明（见微信支付账单导出）
    "alipay": 4,    # 支付宝 ~4 行说明
    "generic": 0,   # 任意银行 / 平台 CSV：动态找表头，不预跳过
}

# 各模板的友好说明（前端显示给用户）
TEMPLATE_DESCRIPTIONS = {
    "wechat": "微信账单 CSV — 支付 / 服务 → 钱包 → 账单 → 申请账单 → 用做记账。",
    "alipay": "支付宝账单 CSV — 我的 → 账单 → 右上角 → 开具交易流水证明 → 邮件接收 CSV。",
    "generic": "任意银行 / 平台 CSV — 按列名（日期 / 金额 / 对方 / 备注 / 收/支）智能匹配。",
}


def _decode_import_bytes(raw: bytes) -> str:
    """微信 / 招行常见 GBK；支付宝可能 UTF-8。统一尝试，失败 fallback 到 UTF-8 with replace。"""
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _resolve_column(headers: list, alias_list: list) -> Optional[int]:
    """根据别名找列索引，返回 -1 表示未找到。"""
    lowered = [h.strip().lower() for h in headers]
    for alias in alias_list:
        target = alias.strip().lower()
        for idx, header in enumerate(lowered):
            if header == target or target in header:
                return idx
    return None


def _normalize_amount(value: str) -> tuple[float, str]:
    """把 '¥ 25.00' / '-5.50' / '+12' 拆成 (abs_value, sign_kind)。
    sign_kind: 'income' 表示符号显示是收入；'expense' 表示是支出；'unknown' 不确定。"""
    if value is None:
        return 0.0, "unknown"
    text = str(value).strip().replace(",", "").replace("¥", "").replace(" ", "")
    if not text:
        return 0.0, "unknown"
    sign_kind = "unknown"
    if text.startswith("-"):
        sign_kind = "expense"
        text = text[1:]
    elif text.startswith("+"):
        sign_kind = "income"
        text = text[1:]
    try:
        return abs(float(text)), sign_kind
    except ValueError:
        return 0.0, "unknown"


def _classify_kind(kind_label: str, sign_kind: str) -> str:
    """优先用显式标签判定 income/expense/transfer，否则用金额符号。默认 expense。"""
    text = (kind_label or "").strip().lower()
    if text:
        for token in _KIND_INCOME_TOKENS:
            if token in text:
                return "income"
        for token in _KIND_EXPENSE_TOKENS:
            if token in text:
                return "expense"
        for token in _KIND_TRANSFER_TOKENS:
            if token in text:
                return "transfer"
    if sign_kind == "income":
        return "income"
    if sign_kind == "expense":
        return "expense"
    return "expense"


def _parse_import_date(value: str) -> str:
    """尝试多种常见时间格式 → ISO。失败返回 utc_now_iso()。"""
    if not value:
        return utc_now_iso()
    text = str(value).strip()
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y.%m.%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y%m%d",
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(text, fmt)
            return dt.replace(tzinfo=timezone.utc).isoformat()
        except ValueError:
            continue
    # ISO 8601 兜底
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).isoformat()
    except (ValueError, TypeError):
        return utc_now_iso()


def _suggest_category(merchant: str, note: str, categories: list) -> Optional[dict]:
    """基于 keywords 模糊匹配。"""
    text = f"{merchant} {note}".lower()
    best = None
    best_score = 0
    for cat in categories:
        keywords = cat.get("keywords") or []
        score = 0
        for kw in keywords:
            if not kw:
                continue
            if str(kw).lower() in text:
                score += len(str(kw))
        if score > best_score:
            best_score = score
            best = cat
    return best if best_score > 0 else None


def _suggest_account(merchant: str, note: str, accounts: list) -> Optional[dict]:
    text = f"{merchant} {note}".lower()
    for account in accounts:
        keywords = account.get("keywords") or []
        for kw in keywords:
            if kw and str(kw).lower() in text:
                return account
    return None


def _suggest_counterparty(merchant: str, counterparties: list) -> Optional[dict]:
    if not merchant:
        return None
    text = merchant.lower()
    for cp in counterparties:
        name = (cp.get("name") or "").lower()
        if name and (name in text or text in name):
            return cp
    return None


def parse_import_content(template: str, raw_bytes: bytes) -> dict:
    """统一入口：返回 {transactions: [...]，warnings: [...]，detected_columns: {...}}。"""
    template_key = template if template in _TEMPLATE_SKIP_LINES else "generic"
    skip_lines = _TEMPLATE_SKIP_LINES[template_key]
    text = _decode_import_bytes(raw_bytes)

    # 跳过前导说明行
    lines = text.splitlines()
    if skip_lines and len(lines) > skip_lines:
        # 找第一个像 CSV 表头的行（包含「日期」/「时间」/「金额」其一）
        # 否则就用配置的 skip_lines
        for offset in range(min(skip_lines + 5, len(lines))):
            line = lines[offset]
            if any(token in line for token in ("交易时间", "日期", "金额", "Date", "Amount", "transaction")):
                lines = lines[offset:]
                break
        else:
            lines = lines[skip_lines:]

    if not lines:
        return {"transactions": [], "warnings": ["文件为空或无可解析内容"], "detected_columns": {}}

    sniffer_text = "\n".join(lines[: min(100, len(lines))])
    try:
        dialect = csv.Sniffer().sniff(sniffer_text, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO("\n".join(lines)), dialect=dialect)
    rows = [row for row in reader if any((cell or "").strip() for cell in row)]
    if not rows:
        return {"transactions": [], "warnings": ["未识别到任何数据行"], "detected_columns": {}}

    headers = [h.strip().strip('"').strip("'") for h in rows[0]]
    body = rows[1:]

    detected = {field: _resolve_column(headers, aliases) for field, aliases in _IMPORT_COLUMN_ALIASES.items()}
    warnings = []
    if detected.get("occurredAt") is None:
        warnings.append("未识别到时间列，默认用导入时间。")
    if detected.get("amount") is None:
        warnings.append("未识别到金额列。请使用通用 CSV 模板手动选列。")
        return {"transactions": [], "warnings": warnings, "detected_columns": detected, "headers": headers}

    # 加载 master data 用来做自动匹配建议
    connection = connect_db()
    try:
        categories = list_categories(connection)
        accounts = list_accounts(connection)
        ledger = load_ledger_settings(connection)
    finally:
        connection.close()
    counterparties = ledger.get("counterparties") or []

    transactions = []
    for row in body:
        def cell(field_key: str) -> str:
            idx = detected.get(field_key)
            if idx is None or idx >= len(row):
                return ""
            return (row[idx] or "").strip()

        amount, sign_kind = _normalize_amount(cell("amount"))
        if amount <= 0:
            continue
        kind = _classify_kind(cell("kindLabel"), sign_kind)
        merchant = cell("merchant")
        note = cell("note")
        occurred = _parse_import_date(cell("occurredAt"))

        suggested_category = _suggest_category(merchant, note, categories)
        suggested_account = _suggest_account(merchant, note, accounts)
        suggested_counterparty = _suggest_counterparty(merchant, counterparties)

        transactions.append({
            "title": merchant or "未命名账单",
            "amount": round(amount, 2),
            "kind": kind,
            "occurredAt": occurred,
            "merchant": merchant,
            "note": note,
            "category": suggested_category and {
                "id": suggested_category.get("id"),
                "name": suggested_category.get("name"),
                "tintHex": suggested_category.get("tintHex"),
            },
            "accountName": (suggested_account or {}).get("name") or (accounts[0]["name"] if accounts else ""),
            "counterpartyId": (suggested_counterparty or {}).get("id"),
            "tags": ["导入", template_key],
            "source": f"import-{template_key}",
            "sourceName": template_key,
            "reimbursementStatus": "notApplicable",
        })

    return {
        "transactions": transactions,
        "warnings": warnings,
        "detected_columns": detected,
        "headers": headers,
        "template": template_key,
    }


def commit_import_transactions(transactions: list) -> dict:
    """批量 INSERT。返回 {imported, failed, errors}。"""
    imported = 0
    failed = 0
    errors = []
    connection = connect_db()
    try:
        for index, tx_payload in enumerate(transactions):
            try:
                row = transaction_row_from_payload(
                    tx_payload,
                    now=utc_now_iso(),
                    transaction_id=str(uuid4()),
                )
                connection.execute(
                    """
                    INSERT INTO transactions (
                        id, title, amount, kind, occurred_at, category_json, tags_json,
                        account_name, from_account_name, to_account_name, merchant, project_name,
                        note, reimbursement_status, source, source_name, counterparty_id, invoice_issued, invoice_attachment_id, tax_category, currency, amount_in_base_currency, created_at, updated_at
                    ) VALUES (
                        :id, :title, :amount, :kind, :occurred_at, :category_json, :tags_json,
                        :account_name, :from_account_name, :to_account_name, :merchant, :project_name,
                        :note, :reimbursement_status, :source, :source_name, :counterparty_id, :invoice_issued, :invoice_attachment_id, :tax_category, :currency, :amount_in_base_currency, :created_at, :updated_at
                    )
                    """,
                    row,
                )
                imported += 1
            except Exception as exc:
                failed += 1
                errors.append({"index": index, "error": str(exc)})
        connection.commit()
    finally:
        connection.close()
    return {"imported": imported, "failed": failed, "errors": errors}


def catchup_recurring_rules(force: bool = False) -> int:
    """扫描所有 enabled 且 next_due_at <= today 的规则，逐条生成交易并前移 next_due_at。
    返回实际生成的交易数。
    debounce：30 秒内只跑一次（除非 force=True），避免每次 GET 都全表扫描。"""
    now_dt = datetime.now()
    if not force and _RECURRING_CATCHUP_LAST_RUN[0] is not None:
        elapsed = (now_dt - _RECURRING_CATCHUP_LAST_RUN[0]).total_seconds()
        if elapsed < _RECURRING_CATCHUP_DEBOUNCE_SEC:
            return 0
    _RECURRING_CATCHUP_LAST_RUN[0] = now_dt

    today = now_dt.date().isoformat()
    generated = 0
    connection = connect_db()
    try:
        rules = connection.execute(
            "SELECT * FROM recurring_rules WHERE enabled = 1 AND next_due_at <= ?",
            (today,),
        ).fetchall()
        for rule in rules:
            try:
                template = json.loads(rule["template_payload_json"]) if rule["template_payload_json"] else {}
            except (json.JSONDecodeError, TypeError):
                template = {}
            cursor_date = rule["next_due_at"]
            end_date = rule["end_date"]
            interval_n = int(rule["interval_n"] or 1)
            frequency = rule["frequency"] or "monthly"
            last_run = rule["last_run_at"]
            for _ in range(120):  # 上限 120 次防呆
                if cursor_date > today:
                    break
                if end_date and cursor_date > end_date:
                    break
                # 用 cursor_date + 12:00 作为发生时间
                occurred_at = f"{cursor_date}T12:00:00+00:00"
                tx_payload = dict(template)
                tx_payload["occurredAt"] = occurred_at
                tx_payload["source"] = template.get("source") or "recurring"
                tx_tags = list(template.get("tags") or [])
                if "周期账目" not in tx_tags:
                    tx_tags.append("周期账目")
                tx_payload["tags"] = tx_tags
                row = transaction_row_from_payload(
                    tx_payload,
                    now=utc_now_iso(),
                    transaction_id=str(uuid4()),
                )
                connection.execute(
                    """
                    INSERT INTO transactions (
                        id, title, amount, kind, occurred_at, category_json, tags_json,
                        account_name, from_account_name, to_account_name, merchant, project_name,
                        note, reimbursement_status, source, source_name, counterparty_id, invoice_issued, invoice_attachment_id, tax_category, currency, amount_in_base_currency, created_at, updated_at
                    ) VALUES (
                        :id, :title, :amount, :kind, :occurred_at, :category_json, :tags_json,
                        :account_name, :from_account_name, :to_account_name, :merchant, :project_name,
                        :note, :reimbursement_status, :source, :source_name, :counterparty_id, :invoice_issued, :invoice_attachment_id, :tax_category, :currency, :amount_in_base_currency, :created_at, :updated_at
                    )
                    """,
                    row,
                )
                last_run = occurred_at
                generated += 1
                cursor_date = advance_due_date(cursor_date, frequency, interval_n)
            connection.execute(
                """
                UPDATE recurring_rules
                SET next_due_at = ?, last_run_at = ?, updated_at = ?
                WHERE id = ?
                """,
                (cursor_date, last_run, utc_now_iso(), rule["id"]),
            )
        connection.commit()
    finally:
        connection.close()
    return generated


def ensure_schema() -> None:
    connection = connect_db()
    try:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS transactions (
                id TEXT PRIMARY KEY,
                title TEXT NOT NULL,
                amount REAL NOT NULL,
                kind TEXT NOT NULL,
                occurred_at TEXT NOT NULL,
                category_json TEXT NOT NULL,
                tags_json TEXT NOT NULL,
                account_name TEXT NOT NULL,
                from_account_name TEXT,
                to_account_name TEXT,
                merchant TEXT NOT NULL,
                note TEXT NOT NULL,
                reimbursement_status TEXT NOT NULL,
                source TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS categories (
                id TEXT PRIMARY KEY,
                payload_json TEXT NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS accounts (
                id TEXT PRIMARY KEY,
                payload_json TEXT NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS ledger_settings (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                payload_json TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS attachments (
                id TEXT PRIMARY KEY,
                transaction_id TEXT NOT NULL,
                mime TEXT,
                size_bytes INTEGER NOT NULL DEFAULT 0,
                original_name TEXT,
                stored_path TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_attachments_tx ON attachments(transaction_id);

            CREATE TABLE IF NOT EXISTS recurring_rules (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                template_payload_json TEXT NOT NULL,
                frequency TEXT NOT NULL,
                interval_n INTEGER NOT NULL DEFAULT 1,
                day_of_period INTEGER,
                start_date TEXT NOT NULL,
                end_date TEXT,
                next_due_at TEXT NOT NULL,
                last_run_at TEXT,
                enabled INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_recurring_due ON recurring_rules(enabled, next_due_at);
            """
        )
        ensure_column(connection, "transactions", "from_account_name", "TEXT")
        ensure_column(connection, "transactions", "to_account_name", "TEXT")
        ensure_column(connection, "transactions", "project_name", "TEXT")
        ensure_column(connection, "transactions", "source_name", "TEXT")
        ensure_column(connection, "transactions", "counterparty_id", "TEXT")
        # 2.2 发票追踪
        ensure_column(connection, "transactions", "invoice_issued", "INTEGER NOT NULL DEFAULT 0")
        ensure_column(connection, "transactions", "invoice_attachment_id", "TEXT")
        # 2.3 税务字段（默认 personal）
        ensure_column(connection, "transactions", "tax_category", "TEXT")
        # W3-H 多币种字段化：交易记原币 + 折算到本位币的快照（避免历史汇率漂移）
        ensure_column(connection, "transactions", "currency", "TEXT")
        ensure_column(connection, "transactions", "amount_in_base_currency", "REAL")
        ensure_column(connection, "transactions", "deleted_at", "TEXT")
        ensure_column(connection, "transactions", "deleted_by", "TEXT")
        ensure_column(connection, "transactions", "deletion_reason", "TEXT")
        ensure_column(connection, "transactions", "deletion_operation_id", "TEXT")
        # 报销核销：垫付支出 → 覆盖它的回款收入 id（NULL = 未核销或快捷按钮手动标记）
        ensure_column(connection, "transactions", "reimbursed_by", "TEXT")
        ensure_audit_schema(connection)
        connection.commit()
    finally:
        connection.close()


def ensure_column(connection: sqlite3.Connection, table: str, column_name: str, definition: str) -> None:
    columns = {
        row["name"]
        for row in connection.execute(f"PRAGMA table_info({table})").fetchall()
    }
    if column_name not in columns:
        connection.execute(f"ALTER TABLE {table} ADD COLUMN {column_name} {definition}")


def seed_default_master_data() -> None:
    connection = connect_db()
    try:
        now = utc_now_iso()
        category_count = connection.execute("SELECT COUNT(*) AS count FROM categories").fetchone()["count"]
        if category_count == 0:
            connection.executemany(
                """
                INSERT INTO categories (id, payload_json, sort_order, updated_at)
                VALUES (?, ?, ?, ?)
                """,
                [
                    (
                        item["id"],
                        json.dumps(item, ensure_ascii=False),
                        index,
                        now,
                    )
                    for index, item in enumerate(default_categories())
                ],
            )

        account_count = connection.execute("SELECT COUNT(*) AS count FROM accounts").fetchone()["count"]
        if account_count == 0:
            connection.executemany(
                """
                INSERT INTO accounts (id, payload_json, sort_order, updated_at)
                VALUES (?, ?, ?, ?)
                """,
                [
                    (
                        item["id"],
                        json.dumps(item, ensure_ascii=False),
                        index,
                        now,
                    )
                    for index, item in enumerate(default_accounts())
                ],
            )

        settings_count = connection.execute("SELECT COUNT(*) AS count FROM ledger_settings").fetchone()["count"]
        if settings_count == 0:
            connection.execute(
                """
                INSERT INTO ledger_settings (id, payload_json, updated_at)
                VALUES (1, ?, ?)
                """,
                (
                    json.dumps(default_ledger_settings(), ensure_ascii=False),
                    now,
                ),
            )
        connection.commit()
    finally:
        connection.close()


def list_categories(connection: sqlite3.Connection) -> list:
    rows = connection.execute(
        """
        SELECT payload_json
        FROM categories
        ORDER BY sort_order ASC, id ASC
        """
    ).fetchall()
    return [json.loads(row["payload_json"]) for row in rows]


def current_balance_for_account(
    connection: sqlite3.Connection,
    account_name: str,
    opening_balance: float,
    classification: str = "asset",
) -> float:
    # openingBalance 是账户的"初始基线"。当前余额 = 初始基线 + 全部相关
    # 交易差额（含 source='adjustment' 的余额对账交易）。当用户在设置页
    # 直接改"当前余额"时，PUT /v1/configuration 会自动写入一条 adjustment
    # 交易补差，让黑洞资金在账本里可追溯（见 _handle_put_configuration）。
    #
    # 资产/负债口径分叉：负债账户（信用卡/贷款）的 openingBalance 与 currentBalance
    # 语义是"正数=已欠"。信用卡刷卡(expense, delta 减)应让"已欠"上升、还款(income/转入)
    # 应让"已欠"下降——故负债用 opening - delta。前端净资产(sign=-1)、availableCredit、
    # totalLiabilities、build_adjustment_payload 都按"正数已欠"口径写，唯此生产者需对齐。
    rows = connection.execute(
        """
        SELECT kind, amount, account_name, from_account_name, to_account_name
        FROM transactions
        WHERE deleted_at IS NULL AND (account_name = ? OR from_account_name = ? OR to_account_name = ?)
        """,
        (account_name, account_name, account_name),
    ).fetchall()
    delta = 0.0
    for row in rows:
        from_account_name, to_account_name = normalized_account_pair(row)
        amount = float(row["amount"] or 0)
        if row["kind"] == "income":
            if to_account_name == account_name:
                delta += amount
        elif row["kind"] == "expense":
            if from_account_name == account_name:
                delta -= amount
        elif row["kind"] == "transfer":
            if from_account_name == account_name:
                delta -= amount
            if to_account_name == account_name:
                delta += amount
    if classification == "liability":
        # 负债：正数=已欠。expense(delta 减) → 已欠增；income/转入(delta 增) → 已欠减。
        return opening_balance - delta
    return opening_balance + delta


def list_accounts(connection: sqlite3.Connection) -> list:
    rows = connection.execute(
        """
        SELECT payload_json
        FROM accounts
        ORDER BY sort_order ASC, id ASC
        """
    ).fetchall()
    items = []
    for row in rows:
        payload = json.loads(row["payload_json"])
        opening_balance = float(payload.get("openingBalance", 0.0))
        payload["threshold"] = coerce_float(payload.get("threshold"), 0.0)
        # W3 阈值多档：低 / 中两条警戒线（< low 绿、low-mid 黄、> mid 红，超 threshold 警告）
        zones = payload.get("thresholdZones") if isinstance(payload.get("thresholdZones"), dict) else {}
        low = coerce_float(zones.get("low"), 0.0)
        mid = coerce_float(zones.get("mid"), 0.0)
        # 缺省按 60% / 85% 推断
        if payload["threshold"] > 0 and (low <= 0 or mid <= 0):
            low = low if low > 0 else round(payload["threshold"] * 0.6, 2)
            mid = mid if mid > 0 else round(payload["threshold"] * 0.85, 2)
        payload["thresholdZones"] = {"low": low, "mid": mid}
        payload["ownership"] = infer_account_ownership(payload)
        # 1.5 资产/负债：默认 asset；信用卡类型回退为 liability
        classification = payload.get("classification")
        if classification not in ("asset", "liability"):
            classification = "liability" if payload.get("type") == "creditCard" else "asset"
        payload["classification"] = classification
        credit_limit = coerce_float(payload.get("creditLimit"), 0.0)
        payload["creditLimit"] = credit_limit if classification == "liability" else 0.0
        payload["currentBalance"] = current_balance_for_account(
            connection,
            payload.get("name", ""),
            opening_balance,
            classification,
        )
        # 信用卡可用额度 = 总额度 - 当前已欠
        if classification == "liability" and credit_limit > 0:
            payload["availableCredit"] = round(credit_limit - payload["currentBalance"], 2)
        items.append(payload)
    return items


def account_ownership_map(connection: sqlite3.Connection) -> dict:
    """Return {account_name: ownership} for fast lookups during view filtering."""
    rows = connection.execute("SELECT payload_json FROM accounts").fetchall()
    mapping: dict = {}
    for row in rows:
        try:
            payload = json.loads(row["payload_json"])
        except (TypeError, json.JSONDecodeError):
            continue
        name = str(payload.get("name") or "").strip()
        if not name:
            continue
        mapping[name] = infer_account_ownership(payload)
    return mapping


def transaction_belongs_to_view(
    transaction: dict,
    view: str,
    ownership_map: dict,
) -> bool:
    """Decide whether a transaction should be included for the requested view.

    view = combined  → always True
    view = company   → ANY of the involved accounts is company-owned
    view = personal  → ANY of the involved accounts is personal-owned
    Unspecified-owned accounts are never matched by company/personal views, only
    by combined.
    """
    if view not in {"company", "personal"}:
        return True

    candidates = {
        transaction.get("accountName"),
        transaction.get("fromAccountName"),
        transaction.get("toAccountName"),
        # also check raw row-style keys for in-flight dicts
        transaction.get("account_name"),
        transaction.get("from_account_name"),
        transaction.get("to_account_name"),
    }
    candidates.discard(None)
    candidates.discard("")

    for name in candidates:
        if ownership_map.get(name) == view:
            return True
    return False


def normalize_view_param(value: object) -> str:
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"company", "personal", "combined"}:
            return lowered
    return "combined"


def load_ledger_settings(connection: sqlite3.Connection) -> dict:
    row = connection.execute(
        """
        SELECT payload_json
        FROM ledger_settings
        WHERE id = 1
        """
    ).fetchone()
    if row is None:
        return default_ledger_settings()
    payload = json.loads(row["payload_json"])
    defaults = default_ledger_settings()
    defaults.update({key: value for key, value in payload.items() if key != "projects"})
    defaults["projects"] = normalize_projects(payload.get("projects"))
    defaults["financeSources"] = normalize_finance_sources(payload.get("financeSources"))
    defaults["counterparties"] = normalize_counterparties(payload.get("counterparties"))
    # ledgerMode 迁移：payload 显式带则归一化；老库缺字段 → 按现存账户归属推断，
    # 保证已有经营数据的库不会被误降级成 personal 而藏掉公司维度。
    if "ledgerMode" in payload:
        defaults["ledgerMode"] = normalize_ledger_mode(payload.get("ledgerMode"))
    else:
        defaults["ledgerMode"] = infer_ledger_mode(connection)
    return defaults


def load_configuration_payload() -> dict:
    connection = connect_db()
    try:
        return {
            "categories": list_categories(connection),
            "accounts": list_accounts(connection),
            "settings": load_ledger_settings(connection),
        }
    finally:
        connection.close()


def attachments_for_transactions(connection: sqlite3.Connection, transaction_ids: list) -> dict:
    """Batch-load attachment metadata for many transactions."""
    if not transaction_ids:
        return {}
    placeholders = ",".join("?" * len(transaction_ids))
    rows = connection.execute(
        f"SELECT id, transaction_id, mime, size_bytes, original_name, created_at "
        f"FROM attachments WHERE transaction_id IN ({placeholders}) "
        f"ORDER BY created_at",
        transaction_ids,
    ).fetchall()
    out: dict = {}
    for r in rows:
        out.setdefault(r["transaction_id"], []).append({
            "id": r["id"],
            "mime": r["mime"],
            "sizeBytes": r["size_bytes"],
            "originalName": r["original_name"],
            "createdAt": r["created_at"],
        })
    return out


def _safe_row_get(row, key, default=None):
    """sqlite3.Row 在缺列时抛 IndexError；dict 在缺键时抛 KeyError。统一兜底。"""
    if row is None:
        return default
    if isinstance(row, dict):
        return row.get(key, default)
    try:
        return row[key]
    except (IndexError, KeyError):
        return default


def row_to_transaction(row: sqlite3.Row, attachments: Optional[list] = None) -> dict:
    from_account_name, to_account_name = normalized_account_pair(row)
    return {
        "id": row["id"],
        "title": row["title"],
        "amount": row["amount"],
        "kind": row["kind"],
        "occurredAt": row["occurred_at"],
        "category": json.loads(row["category_json"]),
        "tags": json.loads(row["tags_json"]),
        "accountName": row["account_name"],
        "fromAccountName": from_account_name,
        "toAccountName": to_account_name,
        "merchant": row["merchant"],
        "projectName": row["project_name"],
        "note": row["note"],
        "reimbursementStatus": row["reimbursement_status"],
        "reimbursedBy": _safe_row_get(row, "reimbursed_by"),
        "source": row["source"],
        "sourceName": row["source_name"],
        "counterpartyId": _safe_row_get(row, "counterparty_id"),
        "invoiceIssued": bool(_safe_row_get(row, "invoice_issued", 0)),
        "invoiceAttachmentId": _safe_row_get(row, "invoice_attachment_id"),
        "taxCategory": _safe_row_get(row, "tax_category") or "personal",
        "currency": _safe_row_get(row, "currency"),
        "amountInBaseCurrency": _safe_row_get(row, "amount_in_base_currency"),
        "deletedAt": _safe_row_get(row, "deleted_at"),
        "deletedBy": _safe_row_get(row, "deleted_by"),
        "deletionReason": _safe_row_get(row, "deletion_reason"),
        "deletionOperationId": _safe_row_get(row, "deletion_operation_id"),
        "attachments": attachments or [],
    }


def dict_to_transaction(row: dict, attachments: Optional[list] = None) -> dict:
    from_account_name, to_account_name = normalized_account_pair(row)
    return {
        "id": row["id"],
        "title": row["title"],
        "amount": row["amount"],
        "kind": row["kind"],
        "occurredAt": row["occurred_at"],
        "category": json.loads(row["category_json"]),
        "tags": json.loads(row["tags_json"]),
        "accountName": row["account_name"],
        "fromAccountName": from_account_name,
        "toAccountName": to_account_name,
        "merchant": row["merchant"],
        "projectName": row.get("project_name"),
        "note": row["note"],
        "reimbursementStatus": row["reimbursement_status"],
        "reimbursedBy": row.get("reimbursed_by"),
        "source": row["source"],
        "sourceName": row.get("source_name"),
        "counterpartyId": row.get("counterparty_id"),
        "invoiceIssued": bool(row.get("invoice_issued", 0)),
        "invoiceAttachmentId": row.get("invoice_attachment_id"),
        "taxCategory": row.get("tax_category") or "personal",
        "currency": row.get("currency"),
        "amountInBaseCurrency": row.get("amount_in_base_currency"),
        "deletedAt": row.get("deleted_at"),
        "deletedBy": row.get("deleted_by"),
        "deletionReason": row.get("deletion_reason"),
        "deletionOperationId": row.get("deletion_operation_id"),
        "attachments": attachments or [],
    }


def existing_transaction_value(
    row: Optional[Union[sqlite3.Row, dict]],
    key: str,
    default=None,
):
    if row is None:
        return default
    if isinstance(row, dict):
        return row.get(key, default)
    try:
        return row[key]
    except (IndexError, KeyError):
        return default


def build_adjustment_payload(account_name: str, delta: float, classification: str, occurred_at: str) -> dict:
    """构造一条"余额调整"交易 payload。

    用户在设置页修改账户当前余额时被自动调用，把账实差额写成一条
    可追溯的 income / expense 交易，让黑洞资金在账本和桑基图里可见。

    资产账户：余额↑ → income（找回的钱），↓ → expense（丢失的钱）
    负债账户：余额↑ → expense（发现的额外欠款），↓ → income（账外还款）
    """
    abs_delta = abs(delta)
    if classification == "liability":
        kind = "expense" if delta > 0 else "income"
    else:
        kind = "income" if delta > 0 else "expense"
    return {
        "title": "余额调整",
        "amount": abs_delta,
        "kind": kind,
        "occurredAt": occurred_at,
        "category": {
            "id": "adjustment-blackhole",
            "name": "余额调整",
            "systemImage": "scalemass",
            "tintHex": "#7c7d6e",
            "keywords": [],
            "direction": "支出" if kind == "expense" else "收入",
            "group": "黑洞资金",
            "note": "用户对账时系统自动写入",
        },
        "tags": ["余额调整", "黑洞资金"],
        "accountName": account_name,
        "fromAccountName": account_name if kind == "expense" else None,
        "toAccountName": account_name if kind == "income" else None,
        "merchant": "余额调整",
        "projectName": "",
        "note": "系统记录的账户对账差额（黑洞资金）",
        "reimbursementStatus": "notApplicable",
        "source": "adjustment",
        "sourceName": "余额调整",
    }


def derive_currency_for_account(account_name: str, fallback: str = "CNY") -> str:
    """根据账户名查 account.currency；找不到回退默认。"""
    if not account_name:
        return fallback
    connection = connect_db()
    try:
        rows = connection.execute("SELECT payload_json FROM accounts").fetchall()
    finally:
        connection.close()
    for row in rows:
        try:
            payload = json.loads(row["payload_json"])
        except (json.JSONDecodeError, TypeError):
            continue
        if payload.get("name") == account_name:
            return str(payload.get("currency") or fallback).upper()
    return fallback


def convert_to_base_currency(amount: float, currency: str, base_currency: str, rates: dict) -> float:
    """W3-H：按当前汇率把 amount 折算到 base_currency。
    rates 格式：{ XXX: N }，意思是 1 XXX = N base。"""
    if not currency or currency.upper() == base_currency.upper():
        return amount
    rate = rates.get(currency.upper())
    if not isinstance(rate, (int, float)) or rate <= 0:
        return amount  # 没找到汇率就不折算（数据完整性优先于精度）
    return amount * float(rate)


def transaction_row_from_payload(
    payload: dict,
    *,
    now: Optional[str] = None,
    transaction_id: Optional[str] = None,
    existing_row: Optional[Union[sqlite3.Row, dict]] = None,
) -> dict:
    timestamp = now or utc_now_iso()
    existing_category_json = existing_transaction_value(existing_row, "category_json", "")
    existing_tags_json = existing_transaction_value(existing_row, "tags_json", "")

    try:
        existing_category = json.loads(existing_category_json) if existing_category_json else None
    except json.JSONDecodeError:
        existing_category = None
    try:
        existing_tags = json.loads(existing_tags_json) if existing_tags_json else []
    except json.JSONDecodeError:
        existing_tags = []

    category = payload.get("category") or existing_category or {
        "id": str(uuid4()),
        "name": "未分类",
        "systemImage": "tray",
        "tintHex": "#607D8B",
        "keywords": [],
    }
    tags = payload.get("tags")
    if not isinstance(tags, list):
        tags = existing_tags if isinstance(existing_tags, list) else []

    title = str(
        payload.get("title")
        or existing_transaction_value(existing_row, "title", "")
        or payload.get("merchant")
        or "未命名账单"
    )
    merchant = str(
        payload.get("merchant")
        or existing_transaction_value(existing_row, "merchant", "")
        or title
    )
    kind = str(
        payload.get("type")
        or payload.get("kind")
        or existing_transaction_value(existing_row, "kind", "expense")
    )
    occurred_at = normalize_iso8601(
        payload.get("occurredAt"),
        existing_transaction_value(existing_row, "occurred_at", timestamp),
    )
    account_name = str(
        payload.get("accountName")
        or existing_transaction_value(existing_row, "account_name", "默认账户")
        or "默认账户"
    )
    from_account_name = (
        payload.get("fromAccountName")
        if "fromAccountName" in payload
        else existing_transaction_value(existing_row, "from_account_name")
    )
    to_account_name = (
        payload.get("toAccountName")
        if "toAccountName" in payload
        else existing_transaction_value(existing_row, "to_account_name")
    )
    project_name = (
        payload.get("projectName")
        if "projectName" in payload
        else existing_transaction_value(existing_row, "project_name")
    )
    source_name = (
        payload.get("sourceName")
        if "sourceName" in payload
        else existing_transaction_value(existing_row, "source_name")
    )
    counterparty_id = (
        payload.get("counterpartyId")
        if "counterpartyId" in payload
        else existing_transaction_value(existing_row, "counterparty_id")
    )
    invoice_issued = (
        coerce_bool(payload.get("invoiceIssued"), False)
        if "invoiceIssued" in payload
        else bool(existing_transaction_value(existing_row, "invoice_issued", 0))
    )
    invoice_attachment_id = (
        payload.get("invoiceAttachmentId")
        if "invoiceAttachmentId" in payload
        else existing_transaction_value(existing_row, "invoice_attachment_id")
    )
    tax_category = (
        payload.get("taxCategory")
        if "taxCategory" in payload
        else existing_transaction_value(existing_row, "tax_category")
    )
    if tax_category not in (
        "business-income",
        "business-expense-deductible",
        "business-expense-nondeductible",
        "personal",
        "transfer",
    ):
        tax_category = "personal"

    # W3-H 多币种字段化：写入时计算并存储 currency + amount_in_base_currency 快照
    currency_in = (
        payload.get("currency")
        if "currency" in payload
        else existing_transaction_value(existing_row, "currency")
    )
    amount_in_base_in = (
        payload.get("amountInBaseCurrency")
        if "amountInBaseCurrency" in payload
        else existing_transaction_value(existing_row, "amount_in_base_currency")
    )

    if kind == "income":
        from_account_name = None
        to_account_name = to_account_name or account_name
        account_name = str(to_account_name or account_name)
    elif kind == "expense":
        from_account_name = from_account_name or account_name
        to_account_name = None
        account_name = str(from_account_name or account_name)
    elif kind == "transfer":
        from_account_name = from_account_name or account_name
        to_account_name = to_account_name or payload.get("targetAccountName") or merchant or ""
        account_name = str(from_account_name or account_name)

    return {
        "id": transaction_id or existing_transaction_value(existing_row, "id") or str(uuid4()),
        "title": title,
        "amount": float(payload.get("amount", existing_transaction_value(existing_row, "amount", 0)) or 0),
        "kind": kind,
        "occurred_at": occurred_at,
        "category_json": json.dumps(category, ensure_ascii=False),
        "tags_json": json.dumps(tags, ensure_ascii=False),
        "account_name": account_name,
        "from_account_name": from_account_name,
        "to_account_name": to_account_name,
        "merchant": merchant,
        "project_name": project_name,
        "note": str(payload.get("note", existing_transaction_value(existing_row, "note", "")) or ""),
        "reimbursement_status": str(
            payload.get(
                "reimbursementStatus",
                existing_transaction_value(existing_row, "reimbursement_status", "notApplicable"),
            )
            or "notApplicable"
        ),
        "source": str(payload.get("source", existing_transaction_value(existing_row, "source", "openClaw")) or "openClaw"),
        "source_name": source_name,
        "counterparty_id": counterparty_id or None,
        "invoice_issued": 1 if invoice_issued else 0,
        "invoice_attachment_id": invoice_attachment_id or None,
        "tax_category": tax_category,
        "currency": _resolve_tx_currency_for_save(currency_in, account_name),
        "amount_in_base_currency": _resolve_amount_in_base_for_save(
            amount_in_base_in,
            payload.get("amount", existing_transaction_value(existing_row, "amount", 0)),
            currency_in,
            account_name,
        ),
        "created_at": existing_transaction_value(existing_row, "created_at", timestamp),
        "updated_at": timestamp,
    }


def _resolve_tx_currency_for_save(currency_in: object, account_name: str) -> str:
    """优先用 payload.currency，否则按账户币种推。"""
    if isinstance(currency_in, str) and currency_in.strip():
        return currency_in.strip().upper()
    return derive_currency_for_account(account_name, "CNY")


def _resolve_amount_in_base_for_save(
    explicit: object, amount: object, currency_in: object, account_name: str
) -> float:
    """优先用 payload.amountInBaseCurrency；否则按当前汇率算。"""
    try:
        if explicit is not None:
            return float(explicit)
    except (TypeError, ValueError):
        pass
    amt = coerce_float(amount, 0.0)
    currency = _resolve_tx_currency_for_save(currency_in, account_name)
    # 拉本位币和汇率
    connection = connect_db()
    try:
        settings = load_ledger_settings(connection)
    finally:
        connection.close()
    rates_cfg = settings.get("exchangeRates") or {}
    base = (rates_cfg.get("baseCurrency") or "CNY").upper()
    rates = rates_cfg.get("rates") or {}
    return round(convert_to_base_currency(amt, currency, base, rates), 4)


def normalized_account_pair(row: Union[sqlite3.Row, dict]) -> tuple[Optional[str], Optional[str]]:
    kind = row["kind"]
    account_name = row.get("account_name") if isinstance(row, dict) else row["account_name"]
    from_account_name = row.get("from_account_name") if isinstance(row, dict) else row["from_account_name"]
    to_account_name = row.get("to_account_name") if isinstance(row, dict) else row["to_account_name"]

    if kind == "income":
        return (from_account_name, to_account_name or account_name)
    if kind == "expense":
        return (from_account_name or account_name, to_account_name)
    if kind == "transfer":
        return (from_account_name or account_name, to_account_name)
    return (from_account_name, to_account_name or account_name)


def parse_transaction_datetime(value: str) -> datetime:
    normalized = value.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(normalized)
    except ValueError:
        return datetime.now().astimezone()


def month_key_from_value(value: str) -> str:
    return value[:7]


def month_label_from_key(month_key: str) -> str:
    year, month = month_key.split("-")
    return f"{year[2:]}/{month}"


def format_currency(value: float) -> str:
    rounded = round(float(value or 0))
    return f"¥{rounded:,.0f}"


def month_keys_for_range(count: int) -> list[str]:
    cursor = datetime.now().astimezone().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    keys = []
    for _ in range(count):
        keys.append(f"{cursor.year:04d}-{cursor.month:02d}")
        if cursor.month == 1:
            cursor = cursor.replace(year=cursor.year - 1, month=12)
        else:
            cursor = cursor.replace(month=cursor.month - 1)
    keys.reverse()
    return keys


def category_group_name(transaction: dict) -> str:
    category_raw = transaction.get("category")
    if isinstance(category_raw, dict):
        category_name = category_raw.get("name") or "未分类"
    elif isinstance(category_raw, str):
        category_name = category_raw or "未分类"
    else:
        category_name = "未分类"
    tags = transaction.get("tags") or []
    reimbursement = transaction.get("reimbursementStatus")

    if reimbursement in {"draft", "submitted", "reimbursed"} or any(tag in {"出差", "客户", "办公", "公司"} for tag in tags):
        return "经营支出"
    if category_name in {"住房", "交通"}:
        return "固定成本"
    return "个人支出"


def group_color_map() -> dict:
    return {
        "经营支出": {
            "group": "#3B82F6",
            "children": ["#60A5FA", "#93C5FD", "#BFDBFE", "#DBEAFE"],
        },
        "固定成本": {
            "group": "#10B981",
            "children": ["#34D399", "#6EE7B7", "#A7F3D0", "#D1FAE5"],
        },
        "个人支出": {
            "group": "#F59E0B",
            "children": ["#FBBF24", "#FCD34D", "#FDE68A", "#FEF3C7"],
        },
    }


def metric_change(current: float, previous: float, suffix: str = "%") -> tuple[str, str]:
    if previous == 0:
        if current == 0:
            return ("0.0" + suffix, "up")
        return ("100.0" + suffix, "up")
    delta = ((current - previous) / abs(previous)) * 100
    trend = "up" if delta >= 0 else "down"
    return (f"{delta:+.1f}{suffix}", trend)


def runway_change(current: float, previous: float) -> tuple[str, str]:
    delta = current - previous
    trend = "up" if delta >= 0 else "down"
    return (f"{delta:+.1f}", trend)


def account_bar_color(tint_hex: Optional[str], index: int) -> str:
    palette = [
        "#3B82F6",
        "#60A5FA",
        "#4F46E5",
        "#0EA5E9",
        "#64748B",
        "#22C55E",
    ]
    if tint_hex:
        return tint_hex
    return palette[index % len(palette)]


def current_balance_from_transactions(transactions: list[dict], account_name: str, opening_balance: float = 0.0) -> float:
    delta = opening_balance
    for transaction in transactions:
        amount = float(transaction.get("amount") or 0)
        if transaction.get("kind") == "income":
            if (transaction.get("toAccountName") or transaction.get("accountName")) == account_name:
                delta += amount
        elif transaction.get("kind") == "expense":
            if (transaction.get("fromAccountName") or transaction.get("accountName")) == account_name:
                delta -= amount
        elif transaction.get("kind") == "transfer":
            if (transaction.get("fromAccountName") or transaction.get("accountName")) == account_name:
                delta -= amount
            if transaction.get("toAccountName") == account_name:
                delta += amount
    return delta


def build_dashboard_overview(server_config: dict, view: str = "combined") -> dict:
    view = normalize_view_param(view)
    connection = connect_db()
    try:
        transactions = [
            row_to_transaction(row)
            for row in connection.execute(
                """
                SELECT *
                FROM transactions
                WHERE deleted_at IS NULL
                ORDER BY occurred_at DESC, created_at DESC
                """
            ).fetchall()
        ]
        configuration = {
            "categories": list_categories(connection),
            "accounts": list_accounts(connection),
            "settings": load_ledger_settings(connection),
        }
        ownership_map = account_ownership_map(connection)
    finally:
        connection.close()

    known_account_names = {
        str(account.get("name") or "")
        for account in configuration["accounts"]
        if account.get("name")
    }
    configuration["accounts"] = [
        account for account in configuration["accounts"]
        if not account.get("deletedAt")
    ]

    if view in {"company", "personal"}:
        transactions = [
            tx for tx in transactions
            if transaction_belongs_to_view(tx, view, ownership_map)
        ]
        configuration["accounts"] = [
            account for account in configuration["accounts"]
            if ownership_map.get(account.get("name") or "") == view
        ]

    month_keys = month_keys_for_range(12)
    current_month_key = month_keys[-1]
    month_summary = {
        key: {"income": 0.0, "expense": 0.0}
        for key in month_keys
    }
    expense_children_by_group = defaultdict(lambda: defaultdict(float))
    income_links = defaultdict(float)
    expense_links = defaultdict(float)
    transfer_links = defaultdict(float)
    roi_groups = defaultdict(lambda: {"cost": 0.0, "revenue": 0.0})
    categories_from_transactions = set()
    transaction_accounts = set()

    for transaction in transactions:
        transaction_type = transaction["kind"]
        amount = float(transaction.get("amount") or 0)
        month_key = month_key_from_value(transaction["occurredAt"])
        category_raw = transaction.get("category")
        if isinstance(category_raw, dict):
            category_name = category_raw.get("name") or "未分类"
        elif isinstance(category_raw, str):
            category_name = category_raw or "未分类"
        else:
            category_name = "未分类"
        categories_from_transactions.add(category_name)
        if transaction.get("accountName"):
            transaction_accounts.add(transaction["accountName"])
        if transaction.get("fromAccountName"):
            transaction_accounts.add(transaction["fromAccountName"])
        if transaction.get("toAccountName"):
            transaction_accounts.add(transaction["toAccountName"])
        from_account_name = transaction.get("fromAccountName")
        to_account_name = transaction.get("toAccountName")

        if month_key in month_summary:
            if transaction_type == "income":
                month_summary[month_key]["income"] += amount
            elif transaction_type == "expense":
                month_summary[month_key]["expense"] += amount

        group_label = (transaction.get("tags") or [category_name])[0] if (transaction.get("tags") or []) else category_name
        if transaction_type == "income":
            roi_groups[group_label]["revenue"] += amount
            if month_key == current_month_key:
                source_node_label = transaction.get("merchant") or transaction.get("title") or "收入"
                target_label = to_account_name or transaction.get("accountName") or "未命名账户"
                income_links[(source_node_label, target_label)] += amount
        elif transaction_type == "expense":
            roi_groups[group_label]["cost"] += amount
            if month_key == current_month_key:
                group_name = category_group_name(transaction)
                expense_children_by_group[group_name][category_name] += amount
                source_node_label = from_account_name or transaction.get("accountName") or "未命名账户"
                expense_links[(source_node_label, category_name)] += amount
        elif transaction_type == "transfer":
            if month_key == current_month_key:
                source_node_label = from_account_name or transaction.get("accountName") or "转出账户"
                target_label = to_account_name or transaction.get("merchant") or transaction.get("note") or "其他账户"
                transfer_links[(source_node_label, target_label)] += amount

    previous_month_key = month_keys[-2] if len(month_keys) > 1 else current_month_key
    current_income = month_summary[current_month_key]["income"]
    current_expense = month_summary[current_month_key]["expense"]
    previous_income = month_summary[previous_month_key]["income"]
    previous_expense = month_summary[previous_month_key]["expense"]
    current_profit = current_income - current_expense
    previous_profit = previous_income - previous_expense

    account_lookup = {
        account["name"]: dict(account)
        for account in configuration["accounts"]
        if account.get("name")
    }

    for account_name in sorted(
        name for name in transaction_accounts if name and name not in known_account_names
    ):
        account_lookup[account_name] = {
            "id": f"derived-{account_name}",
            "name": account_name,
            "type": "other",
            "currency": configuration["settings"].get("defaultCurrency", "CNY"),
            "openingBalance": 0.0,
            "currentBalance": current_balance_from_transactions(transactions, account_name, 0.0),
            "tintHex": None,
            "keywords": [],
        }

    def _net_contribution(item: dict) -> float:
        # 净值贡献：资产 = +当前余额；负债(信用卡/贷款) = -已欠。
        # 负债 currentBalance 已是"正数已欠"（见 current_balance_for_account），故取负。
        raw = float(item.get("currentBalance") or item.get("openingBalance") or 0)
        cls = item.get("classification") or ("liability" if item.get("type") == "creditCard" else "asset")
        return -raw if cls == "liability" else raw

    accounts = []
    total_assets = 0.0
    balance_values = []
    sorted_account_items = sorted(
        account_lookup.values(),
        key=_net_contribution,
        reverse=True,
    )
    for index, account in enumerate(sorted_account_items):
        # 汇总与柱状图用净值贡献：负债显示为负条、不再被当作正资产计入 runway。
        contribution = _net_contribution(account)
        total_assets += contribution
        balance_values.append(contribution)
        accounts.append(
            {
                "name": account.get("name") or "未命名账户",
                "amount": round(contribution, 2),
                "color": account_bar_color(account.get("tintHex"), index),
            }
        )

    average_recent_expense = (
        sum(month_summary[key]["expense"] for key in month_keys[-6:]) / max(len(month_keys[-6:]), 1)
    )
    current_runway = total_assets / average_recent_expense if average_recent_expense else 0.0
    previous_expense_window = month_keys[-7:-1] if len(month_keys) >= 7 else month_keys[:-1]
    previous_average_expense = (
        sum(month_summary[key]["expense"] for key in previous_expense_window) / max(len(previous_expense_window), 1)
        if previous_expense_window
        else average_recent_expense
    )
    previous_runway = total_assets / previous_average_expense if previous_average_expense else current_runway
    opex_rate = (current_expense / current_income) * 100 if current_income else 0.0
    previous_opex_rate = (previous_expense / previous_income) * 100 if previous_income else opex_rate

    cashflow_change, cashflow_trend = metric_change(current_income + current_profit, previous_income + previous_profit)
    profit_change, profit_trend = metric_change(current_profit, previous_profit)
    opex_change, opex_trend = metric_change(opex_rate, previous_opex_rate)
    runway_delta, runway_trend = runway_change(current_runway, previous_runway)

    colors = group_color_map()
    sunburst_data = []
    for group_name in ["经营支出", "固定成本", "个人支出"]:
        children = expense_children_by_group.get(group_name, {})
        if not children:
            continue
        palette = colors[group_name]["children"]
        sunburst_data.append(
            {
                "name": group_name,
                "itemStyle": {"color": colors[group_name]["group"]},
                "children": [
                    {
                        "name": category_name,
                        "value": round(amount, 2),
                        "itemStyle": {"color": palette[index % len(palette)]},
                    }
                    for index, (category_name, amount) in enumerate(
                        sorted(children.items(), key=lambda item: item[1], reverse=True)
                    )
                ],
            }
        )

    roi_candidates = sorted(
        roi_groups.items(),
        key=lambda item: item[1]["cost"] + item[1]["revenue"],
        reverse=True,
    )[:5]
    roi_data = {
        "projects": [item[0] for item in roi_candidates] or ["暂无数据"],
        "cost": [round(item[1]["cost"], 2) for item in roi_candidates] or [0],
        "revenue": [round(item[1]["revenue"], 2) for item in roi_candidates] or [0],
    }

    sankey_nodes = []
    seen_nodes = set()
    for source, target in list(income_links.keys()) + list(expense_links.keys()) + list(transfer_links.keys()):
        if source and source not in seen_nodes:
            sankey_nodes.append({"name": source, "itemStyle": {"color": "#10B981" if source not in account_lookup else "#3B82F6"}})
            seen_nodes.add(source)
        if target and target not in seen_nodes:
            color = "#EF4444"
            if target in account_lookup:
                color = "#3B82F6"
            sankey_nodes.append({"name": target, "itemStyle": {"color": color}})
            seen_nodes.add(target)

    sankey_links = [
        {"source": source, "target": target, "value": round(value, 2)}
        for mapping in (income_links, transfer_links, expense_links)
        for (source, target), value in mapping.items()
        if source and target and value > 0
    ]

    categories = sorted(
        {category.get("name") for category in configuration["categories"] if category.get("name")} | categories_from_transactions
    )

    ui_transactions = []
    for transaction in transactions:
        transaction_type = transaction["kind"]
        from_account_name = transaction.get("fromAccountName")
        to_account_name = transaction.get("toAccountName")
        merchant = transaction.get("merchant") or transaction.get("title") or ""
        category_name = (transaction.get("category") or {}).get("name") or ""
        primary_tag = (transaction.get("tags") or [None])[0]

        if transaction_type == "income":
            from_label = merchant
            to_label = to_account_name or transaction.get("accountName") or "未命名账户"
            type_title = "收入"
        elif transaction_type == "expense":
            from_label = from_account_name or transaction.get("accountName") or "未命名账户"
            to_label = merchant or category_name or "支出"
            type_title = "支出"
        else:
            from_label = from_account_name or transaction.get("accountName") or "转出账户"
            to_label = to_account_name or merchant or transaction.get("note") or "转入账户"
            type_title = "转账"

        note_parts = [part for part in [transaction.get("note"), source_label(transaction), reimbursement_note(transaction)] if part]
        ui_transactions.append(
            {
                "id": transaction["id"],
                "date": transaction["occurredAt"][:10],
                "type": type_title,
                "amount": round(float(transaction.get("amount") or 0), 2) * (-1 if transaction_type == "expense" else 1),
                "from": from_label,
                "to": to_label,
                "note": " · ".join(dict.fromkeys(note_parts)),
                "category": category_name or None,
                "project": primary_tag,
                "tags": transaction.get("tags") or [],
                "accountName": transaction.get("accountName"),
                "merchant": merchant,
                "reimbursementStatus": transaction.get("reimbursementStatus"),
                "source": transaction.get("source"),
            }
        )

    suggestion = {
        "title": "智能财务建议",
        "description": (
            f"本月净现金流 {format_currency(current_profit)}。建议继续通过 OpenClaw 统一记账，并定期在手机端校准账户当前余额，网页与 iPhone 看板会同步更准确。"
        ),
        "actionLabel": "刷新看板",
    }

    return {
        "health": {
            "nodeName": server_config["nodeName"],
            "status": "ok",
            "openClawConnected": True,
            "remoteAccess": server_config.get("remoteAccess", "局域网"),
            "version": "0.2.0",
            "lastIngestedAt": server_config.get("lastIngestedAt"),
        },
        "dashboard": {
            "kpis": {
                "currentCashFlow": {"value": round(total_assets, 2), "display": format_currency(total_assets), "change": cashflow_change, "trend": cashflow_trend},
                "monthlyNetProfit": {"value": round(current_profit, 2), "display": format_currency(current_profit), "change": profit_change, "trend": profit_trend},
                "opexRate": {"value": round(opex_rate, 2), "display": f"{opex_rate:.1f}%", "change": opex_change, "trend": opex_trend},
                "emergencyRunway": {"value": round(current_runway, 2), "display": f"{current_runway:.1f} 个月", "change": runway_delta, "trend": runway_trend},
            },
            "trendData": {
                "months": [month_label_from_key(key) for key in month_keys],
                "income": [round(month_summary[key]["income"], 2) for key in month_keys],
                "expense": [round(month_summary[key]["expense"], 2) for key in month_keys],
            },
            "sunburstData": sunburst_data,
            "roiData": roi_data,
            "sankeyData": {
                "nodes": sankey_nodes,
                "links": sankey_links,
            },
            "accounts": accounts,
            "suggestion": suggestion,
        },
        "allTransactions": ui_transactions,
        "categories": categories,
        "meta": {
            "month": current_month_key,
            "transactionCount": len(ui_transactions),
            "rawTransactionCount": len(transactions),
            "bookMode": configuration["settings"].get("bookMode"),
        },
    }


def source_label(transaction: dict) -> Optional[str]:
    source = transaction.get("source")
    if not source:
        return None
    return {
        "openClaw": "OpenClaw",
        "manual": "手动录入",
        "imported": "导入",
        "localAPI": "本地 API",
        "openCrow": "OpenCrow",
    }.get(source, source)


def reimbursement_note(transaction: dict) -> Optional[str]:
    status = transaction.get("reimbursementStatus")
    if not status or status == "notApplicable":
        return None
    return {
        "draft": "待报销",
        "submitted": "报销中",
        "reimbursed": "已报销",
        "rejected": "已驳回",
    }.get(status, status)


# ----------------------------------------------------------------------------
# Excel export
# ----------------------------------------------------------------------------

EXPORT_KIND_LABEL = {
    "income": "收入",
    "expense": "支出",
    "transfer": "转账",
}

EXPORT_REIMBURSEMENT_LABEL = {
    "draft": "待报销",
    "submitted": "报销中",
    "reimbursed": "已报销",
    "rejected": "已驳回",
    "notApplicable": "—",
}

EXPORT_VIEW_LABEL = {
    "company": "公司视角",
    "personal": "个人视角",
    "combined": "合并视角",
}


def _within_range(occurred_at: str, from_date: Optional[str], to_date: Optional[str]) -> bool:
    if not isinstance(occurred_at, str) or len(occurred_at) < 10:
        return True
    day = occurred_at[:10]
    if from_date and day < from_date:
        return False
    if to_date and day > to_date:
        return False
    return True


def build_tax_report_workbook(year: int, quarter: Optional[int] = None) -> bytes:
    """2.3 报税导出：5 sheet Excel — 业务收入 / 可抵扣 / 不可抵扣 / 季度汇总 / 说明。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required — run `pip3 install openpyxl`.") from exc

    if quarter and quarter in (1, 2, 3, 4):
        start_month = (quarter - 1) * 3 + 1
        end_month = start_month + 2
        from_date = f"{year:04d}-{start_month:02d}-01"
        # 最后一日：end_month 末
        from calendar import monthrange
        end_day = monthrange(year, end_month)[1]
        to_date = f"{year:04d}-{end_month:02d}-{end_day:02d}"
        period_label = f"{year} 年 Q{quarter}"
    else:
        from_date = f"{year:04d}-01-01"
        to_date = f"{year:04d}-12-31"
        period_label = f"{year} 年全年"

    connection = connect_db()
    try:
        rows = connection.execute(
            "SELECT * FROM transactions WHERE deleted_at IS NULL ORDER BY occurred_at ASC, created_at ASC"
        ).fetchall()
        ledger = load_ledger_settings(connection)
    finally:
        connection.close()

    tax_config = ledger.get("taxConfig") or default_tax_config()
    transactions = [row_to_transaction(row) for row in rows]
    transactions = [
        tx for tx in transactions if _within_range(tx.get("occurredAt") or "", from_date, to_date)
    ]

    by_tax: dict[str, list] = defaultdict(list)
    for tx in transactions:
        by_tax[tx.get("taxCategory") or "personal"].append(tx)

    business_income = by_tax.get("business-income", [])
    deductible = by_tax.get("business-expense-deductible", [])
    nondeductible = by_tax.get("business-expense-nondeductible", [])

    income_total = sum(tx["amount"] for tx in business_income)
    deductible_total = sum(tx["amount"] for tx in deductible)
    nondeductible_total = sum(tx["amount"] for tx in nondeductible)
    profit = income_total - deductible_total
    vat_rate = float(tax_config.get("vatRate", 0.03))
    personal_threshold = float(tax_config.get("personalThreshold", 60000))
    personal_rate = float(tax_config.get("personalRate", 0.20))
    seb_rate = float(tax_config.get("sebRate", 0.10))

    # 简化预估
    vat_estimate = max(income_total, 0) * vat_rate
    taxable_profit = max(profit - personal_threshold, 0)
    personal_tax_estimate = taxable_profit * personal_rate
    seb_estimate = max(profit, 0) * seb_rate

    workbook = Workbook()
    header_font = Font(name="PingFang SC", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="CC785C")
    sub_header_font = Font(name="PingFang SC", bold=True, color="141413", size=11)
    sub_header_fill = PatternFill("solid", fgColor="EFE9DE")
    money_format = '_-¥* #,##0.00_-;-¥* #,##0.00;_-¥* "-"_-'

    def style_row(sheet, row_idx: int, font, fill):
        for cell in sheet[row_idx]:
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def write_tx_sheet(sheet, items, label):
        sheet.append(["日期", "类型", "标题", "金额", "账户", "对方 / 商户", "项目", "备注"])
        style_row(sheet, 1, header_font, header_fill)
        for tx in items:
            sheet.append([
                (tx.get("occurredAt") or "")[:10],
                EXPORT_KIND_LABEL.get(tx.get("kind") or "", tx.get("kind")),
                tx.get("title"),
                float(tx.get("amount") or 0),
                tx.get("accountName"),
                tx.get("merchant"),
                tx.get("projectName") or "",
                tx.get("note") or "",
            ])
        # 金额格式
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = money_format
        for col_idx, width in enumerate([12, 8, 30, 14, 16, 24, 18, 32], start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        # 合计行
        if items:
            sheet.append(["", "", "合计", sum(t["amount"] for t in items), "", "", "", ""])
            style_row(sheet, sheet.max_row, sub_header_font, sub_header_fill)
            sheet.cell(row=sheet.max_row, column=4).number_format = money_format

    sheet1 = workbook.active
    sheet1.title = "1.业务收入"
    write_tx_sheet(sheet1, business_income, "业务收入")

    sheet2 = workbook.create_sheet("2.可抵扣支出")
    write_tx_sheet(sheet2, deductible, "可抵扣支出")

    sheet3 = workbook.create_sheet("3.不可抵扣")
    write_tx_sheet(sheet3, nondeductible, "不可抵扣")

    # Sheet 4: 汇总
    summary = workbook.create_sheet("4.汇总")
    summary.append([f"{period_label} · 报税汇总（仅供参考）"])
    style_row(summary, 1, header_font, header_fill)
    summary.append([])
    rows_data = [
        ("业务收入合计", income_total),
        ("可抵扣支出合计", deductible_total),
        ("不可抵扣支出合计", nondeductible_total),
        ("净利润（收入 − 可抵扣）", profit),
        ("", None),
        (f"预估增值税（按 {vat_rate * 100:.2f}%）", vat_estimate),
        (f"应税利润（净利润 − 起征点 {personal_threshold:.0f}）", taxable_profit),
        (f"预估个税（按 {personal_rate * 100:.2f}%）", personal_tax_estimate),
        (f"预估社保 / 公积金（按 {seb_rate * 100:.2f}%）", seb_estimate),
    ]
    for label, val in rows_data:
        if val is None:
            summary.append([label])
        else:
            summary.append([label, float(val)])
            summary.cell(row=summary.max_row, column=2).number_format = money_format
    summary.column_dimensions["A"].width = 36
    summary.column_dimensions["B"].width = 18

    # Sheet 5: 说明
    note_sheet = workbook.create_sheet("5.说明")
    note_sheet.append(["字段说明 / 注意事项"])
    style_row(note_sheet, 1, header_font, header_fill)
    note_sheet.append([])
    note_lines = [
        "本数据由本地财务管理系统按交易的「税务分类」字段聚合生成，仅供参考。",
        f"统计区间：{from_date} 至 {to_date}",
        f"汇率 / 单位：{tax_config.get('currency', 'CNY')}",
        "",
        "税务分类映射：",
        "  · business-income → 1.业务收入",
        "  · business-expense-deductible → 2.可抵扣支出",
        "  · business-expense-nondeductible → 3.不可抵扣",
        "  · personal / transfer → 不参与本表",
        "",
        "预估算法（简化版，与真实税法可能有差异）：",
        f"  增值税 ≈ 业务收入合计 × {vat_rate * 100:.2f}%",
        f"  个税  ≈ 应税利润 × {personal_rate * 100:.2f}%",
        f"  社保  ≈ 净利润 × {seb_rate * 100:.2f}%",
        "",
        "请以专业税务人员意见与税务局核定数据为准。",
    ]
    for line in note_lines:
        note_sheet.append([line])
    note_sheet.column_dimensions["A"].width = 80

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_export_workbook(
    view: str = "combined",
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
) -> bytes:
    """Build a 4-sheet xlsx workbook for the given view + date range.

    Sheets: 明细 / 月度汇总 / 分类汇总 / 账户汇总.
    Returns the workbook as raw bytes.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required for /v1/export/xlsx — run `pip3 install openpyxl`."
        ) from exc

    view = normalize_view_param(view)

    connection = connect_db()
    try:
        rows = connection.execute(
            """
            SELECT *
            FROM transactions
            WHERE deleted_at IS NULL
            ORDER BY occurred_at DESC, created_at DESC
            """
        ).fetchall()
        accounts = list_accounts(connection)
        categories = list_categories(connection)
        ownership_map = account_ownership_map(connection)
    finally:
        connection.close()

    transactions = [row_to_transaction(row) for row in rows]
    if view in {"company", "personal"}:
        transactions = [
            tx for tx in transactions
            if transaction_belongs_to_view(tx, view, ownership_map)
        ]

    if from_date or to_date:
        transactions = [
            tx for tx in transactions
            if _within_range(tx.get("occurredAt") or "", from_date, to_date)
        ]

    workbook = Workbook()

    header_font = Font(name="PingFang SC", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="CC785C")
    sub_header_font = Font(name="PingFang SC", bold=True, color="141413", size=11)
    sub_header_fill = PatternFill("solid", fgColor="EFE9DE")
    body_font = Font(name="PingFang SC", size=10)
    money_format = '_-¥* #,##0.00_-;-¥* #,##0.00;_-¥* "-"_-'

    def style_header(sheet, row_idx: int) -> None:
        for cell in sheet[row_idx]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def style_sub_header(sheet, row_idx: int) -> None:
        for cell in sheet[row_idx]:
            cell.font = sub_header_font
            cell.fill = sub_header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # ---- Sheet 1: 明细 ----
    detail_sheet = workbook.active
    detail_sheet.title = "明细"
    detail_headers = [
        "日期", "时间", "类型", "标题", "金额", "主账户",
        "出账账户", "入账账户", "分类", "分组", "项目",
        "标签", "商户", "备注", "报销状态", "来源",
    ]
    detail_sheet.append(detail_headers)
    style_header(detail_sheet, 1)

    for tx in transactions:
        occurred = tx.get("occurredAt") or ""
        date_part = occurred[:10] if len(occurred) >= 10 else occurred
        time_part = occurred[11:19] if len(occurred) >= 19 else ""
        category = tx.get("category") or {}
        category_name = category.get("name") if isinstance(category, dict) else str(category or "")
        category_group = category.get("group") if isinstance(category, dict) else ""
        amount = float(tx.get("amount") or 0)
        signed_amount = amount
        if tx.get("kind") == "expense":
            signed_amount = -amount
        detail_sheet.append([
            date_part,
            time_part,
            EXPORT_KIND_LABEL.get(tx.get("kind") or "", tx.get("kind") or ""),
            tx.get("title") or "",
            signed_amount,
            tx.get("accountName") or "",
            tx.get("fromAccountName") or "",
            tx.get("toAccountName") or "",
            category_name or "",
            category_group or "",
            tx.get("projectName") or "",
            "、".join(tx.get("tags") or []),
            tx.get("merchant") or "",
            tx.get("note") or "",
            EXPORT_REIMBURSEMENT_LABEL.get(
                tx.get("reimbursementStatus") or "notApplicable",
                tx.get("reimbursementStatus") or "",
            ),
            tx.get("sourceName") or tx.get("source") or "",
        ])

    detail_widths = [12, 10, 8, 26, 12, 14, 14, 14, 12, 14, 12, 18, 18, 28, 12, 12]
    for index, width in enumerate(detail_widths, start=1):
        detail_sheet.column_dimensions[get_column_letter(index)].width = width
    for row_idx in range(2, detail_sheet.max_row + 1):
        cell = detail_sheet.cell(row=row_idx, column=5)
        cell.number_format = money_format
        cell.font = body_font
    detail_sheet.freeze_panes = "A2"

    # ---- Sheet 2: 月度汇总 ----
    month_sheet = workbook.create_sheet("月度汇总")
    month_sheet.append(["月份", "收入", "支出", "结余", "转账总额", "笔数"])
    style_header(month_sheet, 1)

    monthly: dict = defaultdict(lambda: {"income": 0.0, "expense": 0.0, "transfer": 0.0, "count": 0})
    for tx in transactions:
        month_key = (tx.get("occurredAt") or "")[:7] or "未知"
        amount = float(tx.get("amount") or 0)
        kind = tx.get("kind")
        if kind == "income":
            monthly[month_key]["income"] += amount
        elif kind == "expense":
            monthly[month_key]["expense"] += amount
        elif kind == "transfer":
            monthly[month_key]["transfer"] += amount
        monthly[month_key]["count"] += 1

    for key in sorted(monthly.keys(), reverse=True):
        stats = monthly[key]
        month_sheet.append([
            key,
            round(stats["income"], 2),
            round(stats["expense"], 2),
            round(stats["income"] - stats["expense"], 2),
            round(stats["transfer"], 2),
            stats["count"],
        ])

    month_widths = [12, 14, 14, 14, 14, 10]
    for index, width in enumerate(month_widths, start=1):
        month_sheet.column_dimensions[get_column_letter(index)].width = width
    for row_idx in range(2, month_sheet.max_row + 1):
        for col_idx in range(2, 6):
            month_sheet.cell(row=row_idx, column=col_idx).number_format = money_format
    month_sheet.freeze_panes = "A2"

    # ---- Sheet 3: 分类汇总 ----
    category_sheet = workbook.create_sheet("分类汇总")
    category_sheet.append(["分类", "分组", "笔数", "总流入", "总流出", "净额", "平均单笔"])
    style_header(category_sheet, 1)

    cat_stats: dict = defaultdict(lambda: {
        "group": "", "count": 0, "income": 0.0, "expense": 0.0,
    })
    for tx in transactions:
        cat = tx.get("category") or {}
        if isinstance(cat, dict):
            name = cat.get("name") or "未分类"
            group = cat.get("group") or ""
        else:
            name = str(cat or "未分类")
            group = ""
        amount = float(tx.get("amount") or 0)
        cat_stats[name]["group"] = group
        cat_stats[name]["count"] += 1
        if tx.get("kind") == "income":
            cat_stats[name]["income"] += amount
        elif tx.get("kind") == "expense":
            cat_stats[name]["expense"] += amount

    for name in sorted(cat_stats.keys(), key=lambda k: cat_stats[k]["expense"] + cat_stats[k]["income"], reverse=True):
        stats = cat_stats[name]
        net = stats["income"] - stats["expense"]
        avg = (stats["income"] + stats["expense"]) / max(stats["count"], 1)
        category_sheet.append([
            name,
            stats["group"],
            stats["count"],
            round(stats["income"], 2),
            round(stats["expense"], 2),
            round(net, 2),
            round(avg, 2),
        ])
    category_widths = [16, 14, 8, 14, 14, 14, 14]
    for index, width in enumerate(category_widths, start=1):
        category_sheet.column_dimensions[get_column_letter(index)].width = width
    for row_idx in range(2, category_sheet.max_row + 1):
        for col_idx in (4, 5, 6, 7):
            category_sheet.cell(row=row_idx, column=col_idx).number_format = money_format
    category_sheet.freeze_panes = "A2"

    # ---- Sheet 4: 账户汇总 ----
    account_sheet = workbook.create_sheet("账户汇总")
    account_sheet.append([
        "账户", "归属", "类型", "期初余额", "流入合计", "流出合计", "期末余额", "笔数",
    ])
    style_header(account_sheet, 1)

    visible_accounts = accounts
    if view in {"company", "personal"}:
        visible_accounts = [
            account for account in accounts
            if ownership_map.get(account.get("name") or "") == view
        ]

    account_flow: dict = defaultdict(lambda: {"in": 0.0, "out": 0.0, "count": 0})
    for tx in transactions:
        amount = abs(float(tx.get("amount") or 0))
        kind = tx.get("kind")
        from_name = tx.get("fromAccountName")
        to_name = tx.get("toAccountName")
        if kind == "income" and to_name:
            account_flow[to_name]["in"] += amount
            account_flow[to_name]["count"] += 1
        elif kind == "expense" and from_name:
            account_flow[from_name]["out"] += amount
            account_flow[from_name]["count"] += 1
        elif kind == "transfer":
            if from_name:
                account_flow[from_name]["out"] += amount
                account_flow[from_name]["count"] += 1
            if to_name:
                account_flow[to_name]["in"] += amount
                account_flow[to_name]["count"] += 1

    ownership_label = {"company": "公司", "personal": "个人", "unspecified": "未指定"}
    for account in visible_accounts:
        name = account.get("name") or ""
        opening = float(account.get("openingBalance") or 0)
        flow = account_flow.get(name, {"in": 0.0, "out": 0.0, "count": 0})
        closing = opening + flow["in"] - flow["out"]
        account_sheet.append([
            name,
            ownership_label.get(account.get("ownership") or "unspecified", "未指定"),
            account.get("uiAccountType") or account.get("type") or "",
            round(opening, 2),
            round(flow["in"], 2),
            round(flow["out"], 2),
            round(closing, 2),
            flow["count"],
        ])
    account_widths = [16, 10, 14, 14, 14, 14, 14, 8]
    for index, width in enumerate(account_widths, start=1):
        account_sheet.column_dimensions[get_column_letter(index)].width = width
    for row_idx in range(2, account_sheet.max_row + 1):
        for col_idx in (4, 5, 6, 7):
            account_sheet.cell(row=row_idx, column=col_idx).number_format = money_format
    account_sheet.freeze_panes = "A2"

    # ---- Metadata sheet (small) ----
    meta_sheet = workbook.create_sheet("说明")
    meta_sheet.append(["项", "值"])
    style_header(meta_sheet, 1)
    meta_sheet.append(["视角", EXPORT_VIEW_LABEL.get(view, view)])
    meta_sheet.append(["起始日期", from_date or "（无限制）"])
    meta_sheet.append(["截止日期", to_date or "（无限制）"])
    meta_sheet.append(["导出时间", utc_now_iso()])
    meta_sheet.append(["明细笔数", len(transactions)])
    meta_sheet.column_dimensions["A"].width = 14
    meta_sheet.column_dimensions["B"].width = 28

    import io
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class FinanceNodeHandler(BaseHTTPRequestHandler):
    server_version = "FinanceNode/0.1.0"

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if self._handle_public_file(parsed.path):
            return

        if not self._authenticate():
            return

        # 1.2 周期性交易：每个认证 GET 都尝试 catchup（debounce 30 秒，cheap）
        try:
            catchup_recurring_rules()
        except Exception as exc:
            sys.stdout.write(f"[recurring] catchup failed: {exc}\n")

        if parsed.path == "/v1/health":
            self._handle_health()
            return
        if parsed.path == "/v1/transactions":
            self._handle_list_transactions(parse_qs(parsed.query))
            return
        if parsed.path == "/v1/recurring":
            self._handle_list_recurring()
            return
        if parsed.path == "/v1/summary/month":
            self._handle_summary_month()
            return
        if parsed.path == "/v1/dashboard/overview":
            self._handle_dashboard_overview()
            return
        if parsed.path == "/v1/budget/status":
            self._handle_budget_status(parse_qs(parsed.query))
            return
        if parsed.path == "/v1/habits":
            self._handle_habits(parse_qs(parsed.query))
            return
        if parsed.path == "/v1/configuration":
            self._handle_get_configuration()
            return
        if parsed.path == "/v1/audit/events":
            self._handle_list_audit_events(parse_qs(parsed.query))
            return
        if parsed.path == "/v1/export/xlsx":
            self._handle_export_xlsx(parse_qs(parsed.query))
            return
        if parsed.path == "/v1/export/tax-report":
            self._handle_export_tax_report(parse_qs(parsed.query))
            return
        # 1.6 附件下载（GET /v1/attachments/{id}）
        parts = [p for p in parsed.path.split("/") if p]
        if len(parts) == 3 and parts[:2] == ["v1", "attachments"]:
            self._handle_get_attachment(parts[2])
            return

        self._send_json(404, {"error": "Not found"})

    def do_POST(self) -> None:
        if not self._authenticate():
            return

        parsed = urlparse(self.path)
        if parsed.path == "/v1/transactions":
            self._handle_create_transaction()
            return
        if parsed.path == "/v1/recurring":
            self._handle_create_recurring()
            return
        if parsed.path == "/v1/import/preview":
            self._handle_import_preview()
            return
        if parsed.path == "/v1/import/commit":
            self._handle_import_commit()
            return
        if parsed.path == "/v1/rates/refresh":
            self._handle_refresh_rates()
            return
        if parsed.path == "/v1/agent/operations":
            self._handle_agent_operation()
            return
        if parsed.path == "/v1/reimbursements/settle":
            self._handle_settle_reimbursement()
            return
        # 1.6 上传附件（POST /v1/transactions/{txid}/attachments）
        parts = [p for p in parsed.path.split("/") if p]
        if len(parts) == 4 and parts[:2] == ["v1", "transactions"] and parts[3] == "attachments":
            self._handle_upload_attachment(parts[2])
            return

        self._send_json(404, {"error": "Not found"})

    def do_PATCH(self) -> None:
        if not self._authenticate():
            return

        parsed = urlparse(self.path)
        parts = [part for part in parsed.path.split("/") if part]
        if len(parts) == 4 and parts[:2] == ["v1", "transactions"] and parts[3] == "reimbursement":
            self._handle_update_reimbursement(parts[2])
            return

        self._send_json(404, {"error": "Not found"})

    def do_PUT(self) -> None:
        if not self._authenticate():
            return

        parsed = urlparse(self.path)
        parts = [part for part in parsed.path.split("/") if part]
        if parsed.path == "/v1/configuration":
            self._handle_put_configuration()
            return
        if len(parts) == 3 and parts[:2] == ["v1", "transactions"]:
            self._handle_update_transaction(parts[2])
            return
        if len(parts) == 3 and parts[:2] == ["v1", "recurring"]:
            self._handle_update_recurring(parts[2])
            return

        self._send_json(404, {"error": "Not found"})

    def do_DELETE(self) -> None:
        if not self._authenticate():
            return

        parsed = urlparse(self.path)
        parts = [part for part in parsed.path.split("/") if part]
        if len(parts) == 3 and parts[:2] == ["v1", "transactions"]:
            self._handle_delete_transaction(parts[2])
            return
        # 1.6 删除附件
        if len(parts) == 3 and parts[:2] == ["v1", "attachments"]:
            self._handle_delete_attachment(parts[2])
            return
        if len(parts) == 3 and parts[:2] == ["v1", "recurring"]:
            self._handle_delete_recurring(parts[2])
            return

        self._send_json(404, {"error": "Not found"})

    def log_message(self, format: str, *args) -> None:
        # 剥离 query string 再落日志：附件直链等会带 ?token=，绝不能进明文日志。
        line = format % args
        if "?" in line:
            line = " ".join(seg.split("?", 1)[0] for seg in line.split(" "))
        sys.stdout.write("%s - - [%s] %s\n" % (self.address_string(), self.log_date_time_string(), line))

    def _handle_public_file(self, path: str) -> bool:
        # 纵深防御：这些静态分支在鉴权之前执行，绝不接受任何含 ".." 的路径段，
        # 否则可穿越到 runtime/ 读取 config.json（明文 token）与整个数据库。
        # 真正的兜底防线是 _send_file 内的包含性校验。
        if any(segment == ".." for segment in path.split("/")):
            self._send_json(404, {"error": "Not found"})
            return True

        if path == "/favicon.ico":
            self.send_response(204)
            self.send_header("Content-Length", "0")
            self.end_headers()
            return True

        if path == "/":
            self.send_response(302)
            self.send_header("Location", "/dashboard/")
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            return True

        if path in {"/dashboard", "/dashboard/"}:
            self._send_file(WEB_DIR / "index.html")
            return True

        if path.startswith("/dashboard/"):
            relative = path.removeprefix("/dashboard/")
            if relative and "." in Path(relative).name:
                target = WEB_DIR / relative
                if target.exists() and target.is_file():
                    self._send_file(target)
                    return True
            self._send_file(WEB_DIR / "index.html")
            return True

        if path.startswith("/"):
            target = WEB_DIR / path.lstrip("/")
            if target.exists() and target.is_file():
                self._send_file(target)
                return True

        return False

    def _authenticate(self) -> bool:
        config = self.server.config
        token = config.get("accessToken", "").strip()
        if not token:
            return True

        authorization = self.headers.get("Authorization", "")
        if authorization.startswith("Bearer ") and hmac.compare_digest(authorization[7:], token):
            return True

        # 附件已改为前端用 Authorization 头拉 Blob（见 api/client.fetchAttachmentBlob），
        # 故彻底移除 ?token= URL 兜底——token 不再出现在任何 URL，杜绝泄漏到历史/日志/Referer。
        self._send_json(401, {"error": "Unauthorized"})
        return False

    def _handle_health(self) -> None:
        config = self.server.config
        response = {
            "nodeName": config["nodeName"],
            "status": "ok",
            "openClawConnected": True,
            "remoteAccess": config["remoteAccess"],
            "version": "0.1.0",
            "lastIngestedAt": config.get("lastIngestedAt"),
        }
        self._send_json(200, response)

    def _handle_list_transactions(self, query: dict) -> None:
        connection = connect_db()
        try:
            rows = connection.execute(
                """
                SELECT *
                FROM transactions
                ORDER BY occurred_at DESC, created_at DESC
                """
            ).fetchall()
            ownership_map = account_ownership_map(connection)
            attachment_map = attachments_for_transactions(connection, [r["id"] for r in rows])
        finally:
            connection.close()

        items = [row_to_transaction(row, attachments=attachment_map.get(row["id"], [])) for row in rows]
        view = normalize_view_param(query.get("view", [None])[0])
        kind = query.get("kind", [None])[0]
        reimbursement_status = query.get("reimbursementStatus", [None])[0]
        tag = query.get("tag", [None])[0]
        search = query.get("q", [None])[0]
        month = query.get("month", [None])[0]
        category_id = query.get("categoryId", [None])[0]
        account_name = query.get("accountName", [None])[0]
        include_deleted = query.get("includeDeleted", ["0"])[0] in {"1", "true", "yes"}
        limit = query.get("limit", [None])[0]

        if not include_deleted:
            items = [item for item in items if not item.get("deletedAt")]

        if view in {"company", "personal"}:
            items = [
                item for item in items
                if transaction_belongs_to_view(item, view, ownership_map)
            ]
        if month and len(month) == 7 and month[4] == "-":
            items = [item for item in items if item["occurredAt"][:7] == month]
        if kind:
            items = [item for item in items if item["kind"] == kind]
        if reimbursement_status:
            items = [item for item in items if item["reimbursementStatus"] == reimbursement_status]
        if tag:
            items = [item for item in items if tag in item["tags"]]
        if category_id:
            items = [item for item in items if item["category"].get("id") == category_id]
        if account_name:
            items = [item for item in items if item["accountName"] == account_name]
        if search:
            needle = search.lower()
            items = [
                item
                for item in items
                if needle in " ".join(
                    [
                        item["title"],
                        item["category"]["name"],
                        item["accountName"],
                        item["merchant"],
                        item.get("projectName") or "",
                        item["note"],
                        item.get("sourceName") or "",
                        " ".join(item["tags"]),
                    ]
                ).lower()
            ]
        if limit:
            try:
                items = items[: max(int(limit), 0)]
            except ValueError:
                pass

        self._send_json(200, items)

    def _handle_summary_month(self) -> None:
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        prefix = query.get("month", [None])[0]
        if not prefix or len(prefix) != 7 or prefix[4] != "-":
            now = datetime.now().astimezone()
            prefix = f"{now.year:04d}-{now.month:02d}"
        view = normalize_view_param(query.get("view", [None])[0])
        connection = connect_db()
        try:
            rows = connection.execute(
                """
                SELECT *
                FROM transactions
                WHERE deleted_at IS NULL AND substr(occurred_at, 1, 7) = ?
                """,
                (prefix,),
            ).fetchall()
            ownership_map = account_ownership_map(connection)
        finally:
            connection.close()

        items = [row_to_transaction(row) for row in rows]
        if view in {"company", "personal"}:
            items = [
                item for item in items
                if transaction_belongs_to_view(item, view, ownership_map)
            ]

        income = sum(item["amount"] for item in items if item["kind"] == "income")
        expense = sum(item["amount"] for item in items if item["kind"] == "expense")
        # 待回款口径须与前端 isPendingReimbursement 对齐：支出且状态 ∈ {draft,submitted,rejected}。
        # 已驳回不是终态（可二次报销/申诉），仍算被欠的钱，不能漏计。
        pending = sum(
            item["amount"]
            for item in items
            if item["kind"] == "expense"
            and item.get("reimbursementStatus") in {"draft", "submitted", "rejected"}
        )
        self._send_json(
            200,
            {
                "month": prefix,
                "view": view,
                "income": income,
                "expense": expense,
                "balance": income - expense,
                "pendingReimbursement": pending,
                "transactionCount": len(items),
            },
        )

    def _handle_dashboard_overview(self) -> None:
        parsed = urlparse(self.path)
        view = normalize_view_param(parse_qs(parsed.query).get("view", [None])[0])
        self._send_json(200, build_dashboard_overview(self.server.config, view=view))

    def _handle_refresh_rates(self) -> None:
        """W3-G 汇率自动拉取：手动触发 → 调外部 API → 写回 settings。"""
        try:
            cfg = refresh_exchange_rates_in_settings()
        except Exception as exc:
            self._send_json(500, {"error": f"刷新失败：{exc}"})
            return
        if cfg.get("lastFetchError"):
            self._send_json(502, {"error": cfg["lastFetchError"], "rates": cfg})
            return
        self._send_json(200, cfg)

    def _handle_import_preview(self) -> None:
        # 1.4 账单导入：JSON body {template, content (base64)}
        payload = self._read_json_body()
        if payload is None:
            return
        template = str(payload.get("template") or "generic")
        b64 = payload.get("content") or ""
        if not isinstance(b64, str) or not b64:
            self._send_json(400, {"error": "Missing or invalid 'content' (base64 string expected)"})
            return
        try:
            raw = base64.b64decode(b64, validate=True)
        except Exception:
            self._send_json(400, {"error": "Invalid base64 content"})
            return
        if len(raw) > 20 * 1024 * 1024:
            self._send_json(413, {"error": "File too large (max 20 MB)"})
            return
        try:
            result = parse_import_content(template, raw)
        except Exception as exc:
            self._send_json(500, {"error": f"Parser failed: {exc}"})
            return
        self._send_json(200, result)

    def _handle_import_commit(self) -> None:
        payload = self._read_json_body()
        if payload is None:
            return
        transactions = payload.get("transactions")
        if not isinstance(transactions, list) or not transactions:
            self._send_json(400, {"error": "Missing or empty 'transactions' list"})
            return
        result = commit_import_transactions(transactions)
        self._send_json(200, result)

    def _handle_list_recurring(self) -> None:
        connection = connect_db()
        try:
            rows = connection.execute(
                "SELECT * FROM recurring_rules ORDER BY enabled DESC, next_due_at ASC, created_at ASC"
            ).fetchall()
        finally:
            connection.close()
        self._send_json(200, [recurring_row_to_dict(row) for row in rows])

    def _handle_create_recurring(self) -> None:
        payload = self._read_json_body()
        if payload is None:
            return
        normalized = normalize_recurring_payload(payload)
        rule_id = payload.get("id") or str(uuid4())
        now = utc_now_iso()
        connection = connect_db()
        try:
            connection.execute(
                """
                INSERT INTO recurring_rules (
                    id, name, template_payload_json, frequency, interval_n, day_of_period,
                    start_date, end_date, next_due_at, last_run_at, enabled, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, ?, ?, ?)
                """,
                (
                    rule_id,
                    normalized["name"],
                    normalized["template_payload_json"],
                    normalized["frequency"],
                    normalized["interval_n"],
                    normalized["day_of_period"],
                    normalized["start_date"],
                    normalized["end_date"],
                    normalized["next_due_at"],
                    normalized["enabled"],
                    now,
                    now,
                ),
            )
            connection.commit()
            row = connection.execute("SELECT * FROM recurring_rules WHERE id = ?", (rule_id,)).fetchone()
        finally:
            connection.close()
        # 立即跑一次 catchup，让 next_due_at 已经过了的规则当场生成
        catchup_recurring_rules(force=True)
        self._send_json(201, recurring_row_to_dict(row))

    def _handle_update_recurring(self, rule_id: str) -> None:
        payload = self._read_json_body()
        if payload is None:
            return
        normalized = normalize_recurring_payload(payload)
        now = utc_now_iso()
        connection = connect_db()
        try:
            existing = connection.execute("SELECT * FROM recurring_rules WHERE id = ?", (rule_id,)).fetchone()
            if not existing:
                self._send_json(404, {"error": "Recurring rule not found"})
                return
            connection.execute(
                """
                UPDATE recurring_rules SET
                    name = ?, template_payload_json = ?, frequency = ?, interval_n = ?,
                    day_of_period = ?, start_date = ?, end_date = ?, next_due_at = ?,
                    enabled = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    normalized["name"],
                    normalized["template_payload_json"],
                    normalized["frequency"],
                    normalized["interval_n"],
                    normalized["day_of_period"],
                    normalized["start_date"],
                    normalized["end_date"],
                    normalized["next_due_at"],
                    normalized["enabled"],
                    now,
                    rule_id,
                ),
            )
            connection.commit()
            row = connection.execute("SELECT * FROM recurring_rules WHERE id = ?", (rule_id,)).fetchone()
        finally:
            connection.close()
        catchup_recurring_rules(force=True)
        self._send_json(200, recurring_row_to_dict(row))

    def _handle_delete_recurring(self, rule_id: str) -> None:
        connection = connect_db()
        try:
            cur = connection.execute("DELETE FROM recurring_rules WHERE id = ?", (rule_id,))
            connection.commit()
            if cur.rowcount == 0:
                self._send_json(404, {"error": "Recurring rule not found"})
                return
        finally:
            connection.close()
        self._send_json(200, {"ok": True, "id": rule_id})

    def _handle_habits(self, query: dict) -> None:
        """记账习惯查询：Agent 记简述账目（如"夜宵 10 元"）前先查这里补全默认值。
        习惯不是独立存储，而是账本的实时投影——按新近度加权统计"该短语/该类别
        历史上最常配的账户/项目/来源/报销状态"。习惯变了（最近记法变了）投影自动跟随；
        单笔例外（如一次商务宴请）压不过日常多数，天然不污染常规默认。"""
        phrase = (query.get("q", [""])[0] or "").strip()
        category = (query.get("category", [""])[0] or "").strip()
        connection = connect_db()
        try:
            rows = connection.execute(
                """
                SELECT title, merchant, note, category_json, account_name, project_name,
                       source_name, reimbursement_status, occurred_at, kind
                FROM transactions
                WHERE deleted_at IS NULL AND source != 'adjustment'
                ORDER BY occurred_at DESC
                LIMIT 500
                """
            ).fetchall()
        finally:
            connection.close()

        now = datetime.now(timezone.utc)

        def recency_weight(occurred_at: str) -> float:
            try:
                dt = datetime.fromisoformat(str(occurred_at).replace("Z", "+00:00"))
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                days = max(0, (now - dt).days)
            except (ValueError, TypeError):
                days = 365
            return 0.98 ** min(days, 365)

        def aggregate(matched: list) -> dict:
            dims = {
                "account": defaultdict(lambda: {"weight": 0.0, "count": 0, "lastUsedAt": ""}),
                "project": defaultdict(lambda: {"weight": 0.0, "count": 0, "lastUsedAt": ""}),
                "source": defaultdict(lambda: {"weight": 0.0, "count": 0, "lastUsedAt": ""}),
                "category": defaultdict(lambda: {"weight": 0.0, "count": 0, "lastUsedAt": ""}),
                "reimbursement": defaultdict(lambda: {"weight": 0.0, "count": 0, "lastUsedAt": ""}),
            }
            for row, weight, category_name in matched:
                values = {
                    "account": str(row["account_name"] or ""),
                    "project": str(row["project_name"] or ""),
                    "source": str(row["source_name"] or ""),
                    "category": category_name,
                    "reimbursement": str(row["reimbursement_status"] or "notApplicable"),
                }
                for dim, value in values.items():
                    if not value:
                        continue
                    slot = dims[dim][value]
                    slot["weight"] += weight
                    slot["count"] += 1
                    slot["lastUsedAt"] = max(slot["lastUsedAt"], str(row["occurred_at"] or ""))
            result = {}
            for dim, buckets in dims.items():
                total = sum(item["weight"] for item in buckets.values()) or 1.0
                ranked = sorted(buckets.items(), key=lambda kv: kv[1]["weight"], reverse=True)[:3]
                result[dim] = [
                    {
                        "value": value,
                        "count": item["count"],
                        "share": round(item["weight"] / total, 3),
                        "lastUsedAt": item["lastUsedAt"],
                    }
                    for value, item in ranked
                ]
            return result

        phrase_matched = []
        category_matched = []
        for row in rows:
            category_name = _category_name_from_row(row)
            text = f"{row['title'] or ''} {row['merchant'] or ''} {row['note'] or ''}"
            weight = recency_weight(row["occurred_at"])
            if phrase and phrase in text:
                phrase_matched.append((row, weight, category_name))
            if category and category_name == category:
                category_matched.append((row, weight, category_name))

        self._send_json(200, {
            "phrase": phrase,
            "category": category,
            "byPhrase": {"matches": len(phrase_matched), **aggregate(phrase_matched)},
            "byCategory": {"matches": len(category_matched), **aggregate(category_matched)},
            "hint": (
                "byPhrase（用户原话词汇）命中时优先采用；其次 byCategory。"
                "share ≥ 0.7 视为强习惯，可按此补全并在回显中标注来源；"
                "share < 0.7 或零命中时向用户确认。单笔例外不会压过日常多数。"
            ),
        })

    def _handle_budget_status(self, query: dict) -> None:
        # 1.1 预算管理：返回每个有月度预算的分类，本月实际花费、剩余、占比。
        prefix = (query.get("month", [None])[0] or "").strip()
        if not prefix or len(prefix) != 7 or prefix[4] != "-":
            now = datetime.now().astimezone()
            prefix = f"{now.year:04d}-{now.month:02d}"

        connection = connect_db()
        try:
            categories = list_categories(connection)
            spend_rows = connection.execute(
                """
                SELECT category_json, amount
                FROM transactions
                WHERE deleted_at IS NULL AND kind = 'expense' AND substr(occurred_at, 1, 7) = ?
                """,
                (prefix,),
            ).fetchall()
        finally:
            connection.close()

        # 历史交易的 category.id 可能是 name（早期数据迁移遗留），所以同时按 id 和 name 累加
        spent_by_id: dict[str, float] = defaultdict(float)
        spent_by_name: dict[str, float] = defaultdict(float)
        for row in spend_rows:
            try:
                category = json.loads(row["category_json"]) if row["category_json"] else {}
            except (json.JSONDecodeError, TypeError):
                category = {}
            amount = float(row["amount"] or 0)
            cid = str(category.get("id") or "").strip()
            cname = str(category.get("name") or "").strip()
            if cid:
                spent_by_id[cid] += amount
            if cname:
                spent_by_name[cname] += amount

        items = []
        for category in categories:
            budget = coerce_float(category.get("monthlyBudget"), 0.0)
            if budget <= 0:
                continue
            cid = str(category.get("id") or "")
            cname = str(category.get("name") or "")
            spent_via_id = spent_by_id.get(cid, 0.0) if cid else 0.0
            spent_via_name = spent_by_name.get(cname, 0.0) if cname else 0.0
            # 当 id != name 时取较大值（历史数据可能只匹配上 name）；当 id == name 时不会重复
            spent = round(max(spent_via_id, spent_via_name), 2)
            remaining = round(budget - spent, 2)
            percent_used = round((spent / budget) * 100, 1) if budget > 0 else 0.0
            items.append(
                {
                    "categoryId": cid,
                    "name": category.get("name") or "未命名分类",
                    "budget": round(budget, 2),
                    "spent": spent,
                    "remaining": remaining,
                    "percentUsed": percent_used,
                    "color": category.get("tintHex") or "#7f91d6",
                }
            )

        items.sort(key=lambda item: item["percentUsed"], reverse=True)

        total_budget = sum(item["budget"] for item in items)
        total_spent = sum(item["spent"] for item in items)
        self._send_json(
            200,
            {
                "month": prefix,
                "items": items,
                "totalBudget": round(total_budget, 2),
                "totalSpent": round(total_spent, 2),
                "totalRemaining": round(total_budget - total_spent, 2),
            },
        )

    def _handle_export_tax_report(self, query: dict) -> None:
        # 2.3 报税导出：?year=YYYY[&quarter=1-4]
        try:
            year = int(query.get("year", [None])[0] or datetime.now().year)
        except (TypeError, ValueError):
            self._send_json(400, {"error": "Invalid 'year' parameter"})
            return
        quarter_raw = query.get("quarter", [None])[0]
        quarter: Optional[int] = None
        if quarter_raw:
            try:
                quarter = int(quarter_raw)
                if quarter not in (1, 2, 3, 4):
                    raise ValueError
            except (TypeError, ValueError):
                self._send_json(400, {"error": "Invalid 'quarter' (must be 1-4)"})
                return

        try:
            workbook_bytes = build_tax_report_workbook(year=year, quarter=quarter)
        except RuntimeError as exc:
            self._send_json(503, {"error": str(exc)})
            return

        suffix = f"{year}" + (f"-Q{quarter}" if quarter else "")
        filename = f"tax_report_{suffix}_{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"

        self.send_response(200)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header("Content-Length", str(len(workbook_bytes)))
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(workbook_bytes)

    def _handle_export_xlsx(self, query: dict) -> None:
        view = normalize_view_param(query.get("view", [None])[0])
        from_date = (query.get("from", [None])[0] or "").strip() or None
        to_date = (query.get("to", [None])[0] or "").strip() or None

        if from_date and (len(from_date) != 10 or from_date[4] != "-" or from_date[7] != "-"):
            self._send_json(400, {"error": "Invalid 'from' date — expected YYYY-MM-DD"})
            return
        if to_date and (len(to_date) != 10 or to_date[4] != "-" or to_date[7] != "-"):
            self._send_json(400, {"error": "Invalid 'to' date — expected YYYY-MM-DD"})
            return

        try:
            workbook_bytes = build_export_workbook(view=view, from_date=from_date, to_date=to_date)
        except RuntimeError as exc:
            self._send_json(503, {"error": str(exc)})
            return

        suffix_parts = ["finance", view]
        if from_date:
            suffix_parts.append(from_date)
        if to_date:
            suffix_parts.append(to_date)
        suffix_parts.append(datetime.now().strftime("%Y%m%d-%H%M%S"))
        filename = "_".join(suffix_parts) + ".xlsx"

        self.send_response(200)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header("Content-Length", str(len(workbook_bytes)))
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(workbook_bytes)

    def _handle_get_configuration(self) -> None:
        self._send_json(200, load_configuration_payload())

    def _handle_list_audit_events(self, query: dict) -> None:
        try:
            limit = min(max(int(query.get("limit", ["100"])[0]), 1), 500)
        except ValueError:
            limit = 100
        connection = connect_db()
        try:
            rows = connection.execute(
                "SELECT * FROM audit_events ORDER BY occurred_at DESC LIMIT ?", (limit,)
            ).fetchall()
        finally:
            connection.close()
        self._send_json(200, [
            {
                "id": row["id"], "occurredAt": row["occurred_at"], "actor": row["actor"],
                "action": row["action"], "entityType": row["entity_type"],
                "entityId": row["entity_id"], "entityName": row["entity_name"],
                "impact": json.loads(row["impact_json"]), "payload": json.loads(row["payload_json"]),
            }
            for row in rows
        ])

    def _handle_agent_operation(self) -> None:
        payload = self._read_json_body()
        if payload is None:
            return
        entity_type = str(payload.get("entityType") or "").strip()
        action = str(payload.get("action") or "").strip()
        entity_id = str(payload.get("id") or "").strip()
        patch = payload.get("data") if isinstance(payload.get("data"), dict) else {}
        actor = str(payload.get("actor") or self.headers.get("X-FinOS-Actor") or "agent").strip()
        if entity_type not in {"account", "category", "financeSource"} or action not in {"create", "update", "delete", "restore"}:
            self._send_json(400, {"error": "Unsupported agent operation"})
            return

        # AI 防护闸：主数据增删改必须携带用户确认原话，随审计事件落库存证。
        user_confirmation = str(payload.get("userConfirmation") or "").strip()
        if not user_confirmation:
            self._send_json(422, {
                "error": "缺少 userConfirmation：AI 不得自行增改主数据，须先向用户提出建议并获得明确同意。",
                "code": "user_confirmation_required",
                "agentInstruction": (
                    "把要做的变更（实体类型、名称、关键字段）讲给用户听，并给出可选方案；"
                    "用户明确同意后重发本请求，userConfirmation 填用户的确认原话。"
                ),
            })
            return

        with MUTATION_LOCK:
            now = utc_now_iso()
            operation_id = str(uuid4())
            connection = connect_db()
            try:
                if entity_type in {"account", "category"}:
                    table = "accounts" if entity_type == "account" else "categories"
                    rows = connection.execute(
                        f"SELECT id, payload_json, sort_order FROM {table} ORDER BY sort_order, id"
                    ).fetchall()
                    records = [dict(json.loads(row["payload_json"]), id=row["id"], _sort=row["sort_order"]) for row in rows]
                else:
                    settings = load_ledger_settings(connection)
                    records = [dict(item, _sort=index) for index, item in enumerate(settings.get("financeSources", []))]

                target = next((item for item in records if item.get("id") == entity_id), None) if entity_id else None
                if action == "create":
                    if not entity_id:
                        entity_id = f"{entity_type}-{uuid4()}"
                    if target:
                        self._send_json(409, {"error": "ID already exists"})
                        return
                    target = {"id": entity_id, **patch, "_sort": len(records)}
                    target.setdefault("name", f"新{entity_type}")
                    if entity_type == "account":
                        target.setdefault("type", "cash")
                        target.setdefault("currency", "CNY")
                        target.setdefault("openingBalance", 0.0)
                        target.setdefault("currentBalance", coerce_float(target["openingBalance"], 0.0))
                        target.setdefault("ownership", "unspecified")
                        target.setdefault("classification", "asset")
                        target.setdefault("systemImage", "wallet.pass")
                        target.setdefault("tintHex", "#607D8B")
                    elif entity_type == "category":
                        target.setdefault("systemImage", "tray")
                        target.setdefault("tintHex", "#607D8B")
                        target.setdefault("keywords", [])
                        target.setdefault("direction", "支出")
                    else:
                        target.setdefault("icon", "wallet.pass")
                    records.append(target)
                elif not target:
                    self._send_json(404, {"error": "Master data not found"})
                    return

                previous = {} if action == "create" else dict(target)
                if action == "update":
                    if target.get("deletedAt"):
                        self._send_json(409, {"error": "Restore deleted master data before updating it"})
                        return
                    target.update(patch)
                elif action == "delete":
                    if target.get("deletedAt"):
                        self._send_json(409, {"error": "Master data is already deleted"})
                        return
                    target["deletedAt"] = now
                    target["deletedBy"] = actor
                    target["deletionReason"] = str(payload.get("reason") or "Agent 删除")
                    if entity_type == "account":
                        opening = coerce_float(target.get("openingBalance"), 0.0)
                        classification = target.get("classification") or ("liability" if target.get("type") == "creditCard" else "asset")
                        balance = current_balance_for_account(connection, str(target.get("name") or ""), opening, classification)
                        target["deletionImpact"] = {
                            "balance": round(balance, 2),
                            "assetDelta": round(-balance, 2) if classification == "asset" else 0.0,
                            "liabilityDelta": round(-balance, 2) if classification == "liability" else 0.0,
                            "netWorthDelta": round(-balance if classification == "asset" else balance, 2),
                        }
                elif action == "restore":
                    for key in ("deletedAt", "deletedBy", "deletionReason", "deletionImpact"):
                        target.pop(key, None)

                target.pop("_sort", None)
                if entity_type in {"account", "category"}:
                    table = "accounts" if entity_type == "account" else "categories"
                    connection.execute(
                        f"UPDATE {table} SET payload_json = ?, updated_at = ? WHERE id = ?",
                        (json.dumps(target, ensure_ascii=False), now, entity_id),
                    ) if action != "create" else connection.execute(
                        f"INSERT INTO {table} (id, payload_json, sort_order, updated_at) VALUES (?, ?, ?, ?)",
                        (entity_id, json.dumps(target, ensure_ascii=False), len(records) - 1, now),
                    )
                else:
                    settings["financeSources"] = [{key: value for key, value in item.items() if key != "_sort"} for item in records]
                    settings["updatedAt"] = now
                    connection.execute(
                        "UPDATE ledger_settings SET payload_json = ?, updated_at = ? WHERE id = 1",
                        (json.dumps(settings, ensure_ascii=False), now),
                    )

                event = {
                    "id": operation_id, "occurredAt": now, "actor": actor, "action": action,
                    "entityType": entity_type, "entityId": entity_id,
                    "entityName": str(target.get("name") or entity_id),
                    "userConfirmation": user_confirmation,
                    "impact": target.get("deletionImpact", {}),
                    "payload": {"before": {k: v for k, v in previous.items() if k != "_sort"}, "after": target},
                }
                append_audit_event(connection, event)
                connection.commit()
            finally:
                connection.close()
            try:
                event["gitCommit"] = checkpoint_database(DB_PATH, RUNTIME_DIR, event)
            except Exception as exc:
                self._send_json(500, {"error": f"Operation saved but Git checkpoint failed: {exc}", "operation": event})
                return
        self._send_json(200, {"ok": True, "operation": event, "entity": target})

    def _handle_put_configuration(self) -> None:
        payload = self._read_json_body()
        if payload is None:
            return

        categories = payload.get("categories") or []
        accounts = payload.get("accounts") or []
        settings = payload.get("settings") or default_ledger_settings()
        now = utc_now_iso()

        # 记账模式：前端会回传 ledgerMode；兜底按提交账户是否含 company 归属推断，
        # 防止旧前端漏传把 dual 库误降级成 personal。
        ledger_mode = normalize_ledger_mode(
            settings.get("ledgerMode"),
            fallback=(
                "dual"
                if any(isinstance(a, dict) and a.get("ownership") == "company" for a in accounts)
                else "personal"
            ),
        )
        # 引用完整性基准：本次提交的全部账户 id 集合。defaultAccountId 不在其中即视为悬空。
        valid_account_ids = {
            str(item.get("id") or f"account-{index + 1}")
            for index, item in enumerate(accounts)
            if isinstance(item, dict)
        }

        normalized_categories = []
        for index, item in enumerate(categories):
            category_id = item.get("id") or f"category-{index + 1}"
            category_name = str(item.get("name") or "未分类").strip() or "未分类"
            normalized_categories.append(
                (
                    category_id,
                    json.dumps(
                        {
                            "id": category_id,
                            "name": category_name,
                            "systemImage": item.get("systemImage", "tray"),
                            "tintHex": item.get("tintHex", "#607D8B"),
                            "keywords": normalize_keywords(item.get("keywords"), category_name),
                            "direction": "收入" if item.get("direction") == "收入" else "支出",
                            "group": str(item.get("group") or ""),
                            "defaultAccountId": scrub_account_ref(item.get("defaultAccountId"), valid_account_ids),
                            "projectId": str(item.get("projectId") or ""),
                            "note": str(item.get("note") or ""),
                            "monthlyBudget": coerce_float(item.get("monthlyBudget"), 0.0),
                            "deletedAt": item.get("deletedAt") or None,
                            "deletedBy": item.get("deletedBy") or None,
                            "deletionReason": item.get("deletionReason") or None,
                        },
                        ensure_ascii=False,
                    ),
                    index,
                    now,
                )
            )

        # 2.6 余额调整审计：先读现存账户，识别哪些是"用户改了当前余额"，
        # 把差额排队为 adjustment 交易，让黑洞资金在账本里可见。
        # openingBalance 字段在前端绑定的是"当前余额输入框"，但对已存在的
        # 账户，其语义是"用户期望的当前余额"——不直接写入 DB 的 openingBalance
        # （那是初始基线），而是写一条补差交易。
        existing_account_payloads: dict = {}
        adjustment_pending: list = []
        # 账户重命名传播：交易按账户名快照引用，改名后必须把历史交易的三列同步改名，
        # 否则历史交易从余额计算中被孤立（[审计:account-rename]）。
        account_renames: list = []
        with closing(connect_db()) as _peek:
            for row in _peek.execute("SELECT id, payload_json FROM accounts").fetchall():
                try:
                    existing_account_payloads[row["id"]] = json.loads(row["payload_json"])
                except (json.JSONDecodeError, TypeError):
                    pass

        normalized_accounts = []
        for index, item in enumerate(accounts):
            account_id = item.get("id") or f"account-{index + 1}"
            account_name = str(item.get("name") or f"账户{index + 1}").strip() or f"账户{index + 1}"
            brand = item.get("brand") or item.get("logoPresetId") or "custom"

            classification_value = (
                item.get("classification")
                if item.get("classification") in ("asset", "liability")
                else ("liability" if item.get("type") == "creditCard" else "asset")
            )
            old_payload = existing_account_payloads.get(account_id)
            if old_payload is not None:
                # 已存在账户：openingBalance 永远保持基线，不受 PUT 影响。
                # 前端"当前余额"输入框绑定 payload.currentBalance；若与
                # 系统已计算出的 currentBalance 不同 → 写一条 adjustment 交易。
                opening_to_persist = float(old_payload.get("openingBalance", 0.0))
                # 若改了名，历史交易此刻仍挂在旧名下——余额必须按旧名查，
                # 否则查新名得 0、误判为"余额被清空"而生成虚假调整交易。
                old_name = str(old_payload.get("name") or "").strip()
                if old_name and old_name != account_name:
                    account_renames.append((old_name, account_name))
                balance_lookup_name = old_name or account_name
                if "currentBalance" in item:
                    intended_current = coerce_float(item.get("currentBalance"), 0.0)
                    with closing(connect_db()) as _calc:
                        current_now = current_balance_for_account(_calc, balance_lookup_name, opening_to_persist, classification_value)
                    delta = intended_current - current_now
                    if abs(delta) > 0.005:
                        adjustment_pending.append({
                            "account_name": account_name,
                            "delta": round(delta, 2),
                            "classification": classification_value,
                        })
            else:
                # 新账户：优先用 currentBalance（即用户在 UI 输入的"当前余额"），
                # 否则回退到 openingBalance / balance 字段
                opening_to_persist = coerce_float(
                    item.get("currentBalance"),
                    coerce_float(item.get("openingBalance"), coerce_float(item.get("balance"), 0.0)),
                )

            normalized_accounts.append(
                (
                    account_id,
                    json.dumps(
                        {
                            "id": account_id,
                            "name": account_name,
                            "type": item.get("type", "other"),
                            "currency": item.get("currency", "CNY"),
                            "openingBalance": opening_to_persist,
                            "brand": brand,
                            "tintHex": item.get("tintHex", "#607D8B"),
                            "symbolName": item.get("symbolName", "creditcard.fill"),
                            "keywords": normalize_keywords(item.get("keywords"), account_name),
                            "uiAccountType": str(item.get("uiAccountType") or ""),
                            "customType": str(item.get("customType") or ""),
                            "logoMode": str(item.get("logoMode") or "preset"),
                            "logoPresetId": str(item.get("logoPresetId") or brand),
                            "logoEmoji": str(item.get("logoEmoji") or ""),
                            "logoImageUrl": str(item.get("logoImageUrl") or ""),
                            "threshold": coerce_float(item.get("threshold"), 0.0),
                            "thresholdZones": (
                                {
                                    "low": coerce_float((item.get("thresholdZones") or {}).get("low"), 0.0),
                                    "mid": coerce_float((item.get("thresholdZones") or {}).get("mid"), 0.0),
                                }
                                if isinstance(item.get("thresholdZones"), dict)
                                else {"low": 0.0, "mid": 0.0}
                            ),
                            "note": str(item.get("note") or ""),
                            "flowRole": str(item.get("flowRole") or ""),
                            # personal 模式下缺省归属一律 personal，避免 infer 把"工资卡"误判成 company；
                            # 已带显式 ownership 的账户（dual 侧数据）仍原样保留，切模式不丢数据。
                            "ownership": normalize_ownership(
                                item.get("ownership"),
                                fallback=(
                                    "personal"
                                    if ledger_mode == "personal"
                                    else infer_account_ownership(item)
                                ),
                            ),
                            # 1.5 资产/负债：缺省 asset；creditLimit 仅 liability 有意义
                            "classification": (
                                item.get("classification")
                                if item.get("classification") in ("asset", "liability")
                                else ("liability" if item.get("type") == "creditCard" else "asset")
                            ),
                            "creditLimit": coerce_float(item.get("creditLimit"), 0.0),
                            "deletedAt": item.get("deletedAt") or None,
                            "deletedBy": item.get("deletedBy") or None,
                            "deletionReason": item.get("deletionReason") or None,
                            "deletionImpact": item.get("deletionImpact") if isinstance(item.get("deletionImpact"), dict) else None,
                        },
                        ensure_ascii=False,
                    ),
                    index,
                    now,
                )
            )

        normalized_settings = {
            "bookMode": settings.get("bookMode", "personalAssistant"),
            "ledgerMode": ledger_mode,
            "defaultCurrency": settings.get("defaultCurrency", "CNY"),
            "baseUnit": settings.get("baseUnit", "yuan"),
            "timezone": settings.get("timezone", "Asia/Shanghai"),
            "allowManualEntry": coerce_bool(settings.get("allowManualEntry"), True),
            "projects": normalize_projects(settings.get("projects")),
            "financeSources": normalize_finance_sources(settings.get("financeSources")),
            "counterparties": normalize_counterparties(settings.get("counterparties")),
            "taxConfig": normalize_tax_config(settings.get("taxConfig")),
            "exchangeRates": normalize_exchange_rates(settings.get("exchangeRates")),
            "updatedAt": now,
        }
        # 引用完整性：清空资金来源 / 对手方里指向不存在账户的 defaultAccountId。
        for _src in normalized_settings["financeSources"]:
            _src["defaultAccountId"] = scrub_account_ref(_src.get("defaultAccountId"), valid_account_ids)
        for _cp in normalized_settings["counterparties"]:
            _cp["defaultAccountId"] = scrub_account_ref(_cp.get("defaultAccountId"), valid_account_ids)

        connection = connect_db()
        try:
            connection.execute("DELETE FROM categories")
            connection.execute("DELETE FROM accounts")
            connection.execute("DELETE FROM ledger_settings")

            if normalized_categories:
                connection.executemany(
                    """
                    INSERT INTO categories (id, payload_json, sort_order, updated_at)
                    VALUES (?, ?, ?, ?)
                    """,
                    normalized_categories,
                )

            if normalized_accounts:
                connection.executemany(
                    """
                    INSERT INTO accounts (id, payload_json, sort_order, updated_at)
                    VALUES (?, ?, ?, ?)
                    """,
                    normalized_accounts,
                )

            connection.execute(
                """
                INSERT INTO ledger_settings (id, payload_json, updated_at)
                VALUES (1, ?, ?)
                """,
                (
                    json.dumps(normalized_settings, ensure_ascii=False),
                    now,
                ),
            )

            # 账户重命名传播：先把历史交易的三列从旧名改到新名，历史交易随账户一起走。
            # 放在余额调整写入之前——调整交易用的是新名，改名后二者一致落到同一账户。
            for old_name, new_name in account_renames:
                connection.execute(
                    "UPDATE transactions SET account_name = ?, updated_at = ? WHERE account_name = ?",
                    (new_name, now, old_name),
                )
                connection.execute(
                    "UPDATE transactions SET from_account_name = ?, updated_at = ? WHERE from_account_name = ?",
                    (new_name, now, old_name),
                )
                connection.execute(
                    "UPDATE transactions SET to_account_name = ?, updated_at = ? WHERE to_account_name = ?",
                    (new_name, now, old_name),
                )

            # 2.6 写入排队的余额调整交易（让黑洞资金可追溯）
            for adj in adjustment_pending:
                adj_payload = build_adjustment_payload(
                    adj["account_name"], adj["delta"], adj["classification"], now
                )
                adj_row = transaction_row_from_payload(
                    adj_payload, now=now, transaction_id=str(uuid4())
                )
                connection.execute(
                    """
                    INSERT INTO transactions (
                        id, title, amount, kind, occurred_at, category_json, tags_json,
                        account_name, from_account_name, to_account_name, merchant, project_name,
                        note, reimbursement_status, source, source_name, counterparty_id, invoice_issued, invoice_attachment_id, tax_category, currency, amount_in_base_currency, created_at, updated_at
                    ) VALUES (
                        :id, :title, :amount, :kind, :occurred_at, :category_json, :tags_json,
                        :account_name, :from_account_name, :to_account_name, :merchant, :project_name,
                        :note, :reimbursement_status, :source, :source_name, :counterparty_id, :invoice_issued, :invoice_attachment_id, :tax_category, :currency, :amount_in_base_currency, :created_at, :updated_at
                    )
                    """,
                    adj_row,
                )

            connection.commit()
        finally:
            connection.close()

        self._send_json(200, load_configuration_payload())

    def _amount_payload_invalid(self, payload: dict) -> bool:
        """校验 payload 里的 amount（若提供）为有限非负数，非法则回 400 并返回 True。"""
        if "amount" not in payload or payload.get("amount") is None:
            return False
        try:
            amt = float(payload["amount"])
        except (TypeError, ValueError):
            self._send_json(400, {"error": "amount 必须是数字"})
            return True
        if amt != amt or amt in (float("inf"), float("-inf")) or amt < 0:
            self._send_json(400, {"error": "amount 必须是有限的非负数"})
            return True
        return False

    def _handle_create_transaction(self) -> None:
        payload = self._read_json_body()
        if payload is None:
            return
        if self._amount_payload_invalid(payload):
            return

        now = utc_now_iso()
        row = transaction_row_from_payload(
            payload,
            now=now,
            transaction_id=payload.get("id") or str(uuid4()),
        )

        connection = connect_db()
        try:
            # AI 防护闸：openClaw 通道引用未知主数据 → 422，逼 Agent 先问用户
            guard = validate_agent_transaction_refs(connection, row)
            if guard is not None:
                self._send_json(422, guard)
                return
            connection.execute(
                """
                INSERT INTO transactions (
                    id, title, amount, kind, occurred_at, category_json, tags_json,
                    account_name, from_account_name, to_account_name, merchant, project_name,
                    note, reimbursement_status, source, source_name,
                    counterparty_id, invoice_issued, invoice_attachment_id,
                    tax_category, currency, amount_in_base_currency,
                    created_at, updated_at
                ) VALUES (
                    :id, :title, :amount, :kind, :occurred_at, :category_json, :tags_json,
                    :account_name, :from_account_name, :to_account_name, :merchant, :project_name,
                    :note, :reimbursement_status, :source, :source_name,
                    :counterparty_id, :invoice_issued, :invoice_attachment_id,
                    :tax_category, :currency, :amount_in_base_currency,
                    :created_at, :updated_at
                )
                """,
                row,
            )
            connection.commit()
        finally:
            connection.close()

        self.server.config["lastIngestedAt"] = now
        self._persist_server_config()
        self._send_json(201, dict_to_transaction(row))

    def _handle_update_transaction(self, transaction_id: str) -> None:
        payload = self._read_json_body()
        if payload is None:
            return
        if self._amount_payload_invalid(payload):
            return

        connection = connect_db()
        try:
            existing_row = connection.execute(
                """
                SELECT *
                FROM transactions
                WHERE id = ?
                """,
                (transaction_id,),
            ).fetchone()
            if existing_row is None:
                self._send_json(404, {"error": "Transaction not found"})
                return

            now = utc_now_iso()
            row = transaction_row_from_payload(
                payload,
                now=now,
                transaction_id=transaction_id,
                existing_row=existing_row,
            )
            # AI 防护闸：仅校验相对原记录变化的引用，避免误伤存量编辑
            guard = validate_agent_transaction_refs(connection, row, existing_row=existing_row)
            if guard is not None:
                self._send_json(422, guard)
                return
            connection.execute(
                """
                UPDATE transactions
                SET title = :title,
                    amount = :amount,
                    kind = :kind,
                    occurred_at = :occurred_at,
                    category_json = :category_json,
                    tags_json = :tags_json,
                    account_name = :account_name,
                    from_account_name = :from_account_name,
                    to_account_name = :to_account_name,
                    merchant = :merchant,
                    project_name = :project_name,
                    note = :note,
                    reimbursement_status = :reimbursement_status,
                    source = :source,
                    source_name = :source_name,
                    counterparty_id = :counterparty_id,
                    invoice_issued = :invoice_issued,
                    invoice_attachment_id = :invoice_attachment_id,
                    tax_category = :tax_category,
                    currency = :currency,
                    amount_in_base_currency = :amount_in_base_currency,
                    updated_at = :updated_at
                WHERE id = :id
                """,
                row,
            )
            connection.commit()
        finally:
            connection.close()

        self.server.config["lastIngestedAt"] = now
        self._persist_server_config()
        self._send_json(200, dict_to_transaction(row))

    def _handle_delete_transaction(self, transaction_id: str) -> None:
        actor = str(self.headers.get("X-FinOS-Actor") or "dashboard").strip()
        reason = str(
            parse_qs(urlparse(self.path).query).get("reason", ["删除流水"])[0] or "删除流水"
        ).strip()
        with MUTATION_LOCK:
            now = utc_now_iso()
            operation_id = str(uuid4())
            connection = connect_db()
            try:
                row = connection.execute("SELECT * FROM transactions WHERE id = ?", (transaction_id,)).fetchone()
                if row is None:
                    self._send_json(404, {"error": "Transaction not found"})
                    return
                transaction = row_to_transaction(row)
                cursor = connection.execute(
                    """
                    UPDATE transactions
                    SET deleted_at = ?, deleted_by = ?, deletion_reason = ?, deletion_operation_id = ?, updated_at = ?
                    WHERE id = ? AND deleted_at IS NULL
                    """,
                    (now, actor, reason, operation_id, now, transaction_id),
                )
                if cursor.rowcount == 0:
                    self._send_json(409, {"error": "Transaction is already deleted"})
                    return
                # 删除的是报销回款收入 → 级联把它名下已核销的垫付退回待报销，
                # 否则留下指向已删除收入的悬空 reimbursed_by（幽灵“已报销”）。
                cascaded = 0
                if transaction.get("kind") == "income":
                    cascade_cursor = connection.execute(
                        """
                        UPDATE transactions
                        SET reimbursement_status = 'draft', reimbursed_by = NULL, updated_at = ?
                        WHERE reimbursed_by = ? AND deleted_at IS NULL
                        """,
                        (now, transaction_id),
                    )
                    cascaded = cascade_cursor.rowcount
                event = {
                    "id": operation_id, "occurredAt": now, "actor": actor, "action": "delete",
                    "entityType": "transaction", "entityId": transaction_id, "entityName": transaction["title"],
                    "impact": {"amount": transaction["amount"], "kind": transaction["kind"], "accountName": transaction["accountName"], "reimbursementsUnsettled": cascaded},
                    "payload": {"before": transaction},
                }
                append_audit_event(connection, event)
                connection.commit()
            finally:
                connection.close()
            try:
                event["gitCommit"] = checkpoint_database(DB_PATH, RUNTIME_DIR, event)
            except Exception as exc:
                self._send_json(500, {"error": f"Transaction deleted but Git checkpoint failed: {exc}", "operation": event})
                return
        self.server.config["lastIngestedAt"] = now
        self._persist_server_config()
        self._send_json(200, {"ok": True, "id": transaction_id, "operation": event})

    def _handle_update_reimbursement(self, transaction_id: str) -> None:
        payload = self._read_json_body()
        if payload is None:
            return

        status = payload.get("status")
        if status not in {"notApplicable", "draft", "submitted", "reimbursed", "rejected"}:
            self._send_json(400, {"error": "Invalid status"})
            return

        connection = connect_db()
        try:
            # 离开 reimbursed 状态时清空回款关联，避免残留悬空的核销链接
            if status == "reimbursed":
                cursor = connection.execute(
                    "UPDATE transactions SET reimbursement_status = ?, updated_at = ? WHERE id = ?",
                    (status, utc_now_iso(), transaction_id),
                )
            else:
                cursor = connection.execute(
                    "UPDATE transactions SET reimbursement_status = ?, reimbursed_by = NULL, updated_at = ? WHERE id = ?",
                    (status, utc_now_iso(), transaction_id),
                )
            connection.commit()
        finally:
            connection.close()

        if cursor.rowcount == 0:
            self._send_json(404, {"error": "Transaction not found"})
            return

        self._send_json(200, {"ok": True, "id": transaction_id, "status": status})

    def _handle_settle_reimbursement(self) -> None:
        """回款核销：一笔报销回款收入 ↔ 多笔垫付支出的批量对账。

        payload: { incomeId, settleIds: [...], unsettleIds: [...] }
        - settleIds: 勾选的垫付 → reimbursed + 挂到 incomeId 名下
        - unsettleIds: 取消勾选的垫付 → 退回 draft；仅允许解开挂在 incomeId 名下的（防串账）
        """
        payload = self._read_json_body()
        if payload is None:
            return

        income_id = str(payload.get("incomeId") or "")
        settle_ids = [str(item) for item in (payload.get("settleIds") or []) if item]
        unsettle_ids = [str(item) for item in (payload.get("unsettleIds") or []) if item]
        if not income_id:
            self._send_json(400, {"error": "Missing incomeId"})
            return
        if not settle_ids and not unsettle_ids:
            self._send_json(400, {"error": "Nothing to settle"})
            return

        connection = connect_db()
        try:
            income_row = connection.execute(
                "SELECT id, kind, deleted_at FROM transactions WHERE id = ?", (income_id,)
            ).fetchone()
            if income_row is None or income_row["deleted_at"]:
                self._send_json(404, {"error": "Income transaction not found"})
                return
            if income_row["kind"] != "income":
                self._send_json(422, {"error": "incomeId must reference an income transaction"})
                return

            now = utc_now_iso()
            settled = 0
            unsettled = 0
            invalid = []
            for expense_id in settle_ids:
                row = connection.execute(
                    "SELECT id, kind, reimbursement_status, deleted_at FROM transactions WHERE id = ?",
                    (expense_id,),
                ).fetchone()
                if (
                    row is None
                    or row["deleted_at"]
                    or row["kind"] != "expense"
                    or row["reimbursement_status"] in (None, "", "notApplicable")
                ):
                    invalid.append(expense_id)
                    continue
                connection.execute(
                    "UPDATE transactions SET reimbursement_status = 'reimbursed', reimbursed_by = ?, updated_at = ? WHERE id = ?",
                    (income_id, now, expense_id),
                )
                settled += 1
            for expense_id in unsettle_ids:
                cursor = connection.execute(
                    "UPDATE transactions SET reimbursement_status = 'draft', reimbursed_by = NULL, updated_at = ? "
                    "WHERE id = ? AND reimbursed_by = ?",
                    (now, expense_id, income_id),
                )
                unsettled += cursor.rowcount
            connection.commit()
        finally:
            connection.close()

        self._send_json(200, {
            "ok": True,
            "incomeId": income_id,
            "settled": settled,
            "unsettled": unsettled,
            "invalid": invalid,
        })

    # ---- 1.6 附件 ----

    def _handle_upload_attachment(self, transaction_id: str) -> None:
        payload = self._read_json_body()
        if payload is None:
            return
        filename = str(payload.get("filename") or "attachment")
        mime = str(payload.get("mime") or "application/octet-stream")
        b64 = payload.get("data") or ""
        if not isinstance(b64, str):
            self._send_json(400, {"error": "Missing or invalid data (base64 string expected)"})
            return
        try:
            raw = base64.b64decode(b64, validate=True)
        except Exception:
            self._send_json(400, {"error": "Invalid base64 data"})
            return
        if len(raw) == 0:
            self._send_json(400, {"error": "Empty file"})
            return
        if len(raw) > ATTACHMENT_MAX_BYTES:
            self._send_json(413, {"error": f"File too large (max {ATTACHMENT_MAX_BYTES // (1024*1024)} MB)"})
            return

        connection = connect_db()
        try:
            tx_row = connection.execute(
                "SELECT id FROM transactions WHERE id = ?", (transaction_id,)
            ).fetchone()
            if tx_row is None:
                self._send_json(404, {"error": "Transaction not found"})
                return

            attachment_id = str(uuid4())
            ext_guess = mimetypes.guess_extension(mime) or ""
            tx_dir = ATTACHMENTS_DIR / transaction_id
            tx_dir.mkdir(parents=True, exist_ok=True)
            stored_filename = f"{attachment_id}{ext_guess}"
            stored_path = tx_dir / stored_filename
            stored_path.write_bytes(raw)

            now = utc_now_iso()
            connection.execute(
                """
                INSERT INTO attachments (id, transaction_id, mime, size_bytes, original_name, stored_path, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (attachment_id, transaction_id, mime, len(raw), filename, str(stored_path), now),
            )
            connection.commit()
        finally:
            connection.close()

        self._send_json(201, {
            "id": attachment_id,
            "transactionId": transaction_id,
            "mime": mime,
            "sizeBytes": len(raw),
            "originalName": filename,
            "createdAt": now,
        })

    def _handle_get_attachment(self, attachment_id: str) -> None:
        connection = connect_db()
        try:
            row = connection.execute(
                "SELECT mime, original_name, stored_path FROM attachments WHERE id = ?",
                (attachment_id,),
            ).fetchone()
        finally:
            connection.close()
        if row is None:
            self._send_json(404, {"error": "Attachment not found"})
            return
        path = Path(row["stored_path"])
        if not path.exists():
            self._send_json(404, {"error": "Stored file missing"})
            return
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", row["mime"] or "application/octet-stream")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "private, max-age=300")
        self.end_headers()
        self.wfile.write(body)

    def _handle_delete_attachment(self, attachment_id: str) -> None:
        connection = connect_db()
        try:
            row = connection.execute(
                "SELECT stored_path FROM attachments WHERE id = ?", (attachment_id,)
            ).fetchone()
            if row is None:
                self._send_json(404, {"error": "Attachment not found"})
                return
            connection.execute("DELETE FROM attachments WHERE id = ?", (attachment_id,))
            connection.commit()
        finally:
            connection.close()
        try:
            Path(row["stored_path"]).unlink(missing_ok=True)
        except Exception:
            pass
        self._send_json(200, {"ok": True, "id": attachment_id})

    def _read_json_body(self) -> Optional[dict]:
        # 读入前先按 Content-Length 卡上限，杜绝超大 body 一次性读入内存耗尽（DoS）。
        # 附件/导入的 base64 内容级 413 校验仍保留；此处是原始请求体的硬顶。
        try:
            length = int(self.headers.get("Content-Length", "0"))
        except ValueError:
            self._send_json(400, {"error": "Invalid Content-Length"})
            return None
        if length < 0 or length > MAX_JSON_BODY_BYTES:
            self._send_json(413, {"error": "Request body too large"})
            return None
        try:
            raw = self.rfile.read(length) if length else b"{}"
            return json.loads(raw.decode("utf-8"))
        except json.JSONDecodeError:
            self._send_json(400, {"error": "Invalid JSON"})
            return None

    def _persist_server_config(self) -> None:
        with CONFIG_PATH.open("w", encoding="utf-8") as file:
            json.dump(self.server.config, file, ensure_ascii=False, indent=2)

    def _send_file(self, path: Path) -> None:
        # 兜底防线：目标解析后必须仍在 WEB_ROOT 内，杜绝目录穿越。
        # 无条件生效（即便未设 accessToken），保护 SSH 私钥、其它库、config.json 等。
        try:
            resolved = path.resolve()
            resolved.relative_to(WEB_ROOT)
        except (ValueError, OSError):
            self._send_json(404, {"error": "Not found"})
            return
        path = resolved

        if not path.exists() or not path.is_file():
            self._send_json(404, {"error": "Not found"})
            return

        body = path.read_bytes()
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        cache_control = "no-cache"
        if "assets" in path.parts or path.suffix in {".woff", ".woff2"}:
            cache_control = "public, max-age=31536000, immutable"
        elif path.suffix == ".png":
            cache_control = "public, max-age=604800"

        accepted_encodings = self.headers.get("Accept-Encoding", "")
        should_gzip = (
            "gzip" in accepted_encodings.lower()
            and len(body) > 1024
            and (
                content_type.startswith("text/")
                or content_type in {
                    "application/javascript",
                    "text/javascript",
                    "application/json",
                    "application/manifest+json",
                    "image/svg+xml",
                }
            )
        )

        if should_gzip:
            body = gzip.compress(body, compresslevel=6)

        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", cache_control)
        if should_gzip:
            self.send_header("Content-Encoding", "gzip")
            self.send_header("Vary", "Accept-Encoding")
        self.end_headers()
        self.wfile.write(body)

    def _send_json(self, status_code: int, payload: Union[dict, list]) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status_code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


class FinanceNodeServer(ThreadingHTTPServer):
    def __init__(self, server_address, request_handler_class, config: dict):
        super().__init__(server_address, request_handler_class)
        self.config = config


def main() -> None:
    config = load_config()
    ensure_schema()
    seed_default_master_data()
    # 1.2 周期性交易：启动时 catchup 一次
    try:
        generated = catchup_recurring_rules(force=True)
        if generated > 0:
            print(f"Recurring catchup generated {generated} transactions", flush=True)
    except Exception as exc:
        print(f"[recurring] startup catchup failed: {exc}", flush=True)

    # W3-G 汇率自动拉取：启动时若 autoFetch=true 拉一次
    try:
        connection = connect_db()
        try:
            settings = load_ledger_settings(connection)
        finally:
            connection.close()
        rates_cfg = settings.get("exchangeRates") or {}
        if rates_cfg.get("autoFetch"):
            cfg = refresh_exchange_rates_in_settings()
            if cfg.get("lastFetchError"):
                print(f"[rates] auto-fetch failed: {cfg['lastFetchError']}", flush=True)
            else:
                print(f"[rates] auto-fetched at {cfg.get('updatedAt')}", flush=True)
    except Exception as exc:
        print(f"[rates] startup fetch failed: {exc}", flush=True)

    host = config.get("host", "127.0.0.1")
    port = int(config.get("port", 31888))
    # 安全告警：绑定非回环地址且未设 token = 财务 API 对网络完全裸奔。
    # 不硬阻断（隧道/可信内网后的高级用法保留），但必须大声提醒。
    if host not in ("127.0.0.1", "localhost", "::1") and not str(config.get("accessToken") or "").strip():
        print(
            f"⚠️  警告：绑定 {host} 且未设置 accessToken —— 财务数据将无鉴权地暴露在网络上！\n"
            f"   请在 runtime/config.json 设置 accessToken，或改回 host=127.0.0.1。",
            file=sys.stderr, flush=True,
        )
    server = FinanceNodeServer((host, port), FinanceNodeHandler, config)
    print(f"Finance Node running on http://{host}:{port}", flush=True)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("Shutting down Finance Node", flush=True)
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
